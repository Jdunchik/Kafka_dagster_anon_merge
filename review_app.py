"""
review_app.py — Центр управления EnigmaSQL (Streamlit).

Полный контроль над пайплайном из браузера:
  • Запуск — собрать, категоризировать, склеить названия, объединить, скачать результат
  • Категории — управление целевыми категориями, классификация, ручные правки
  • Мэтчинг — склейка написаний одного товара в canonical_name
  • Файлы — загрузка/просмотр/удаление исходников и результатов
  • Состояние — просмотр и правка всех JSON-файлов состояния, очистка кэшей

Запуск:
    streamlit run review_app.py
"""
import io
import json
import shutil
import sys
import urllib.request
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import streamlit as st

try:
    import altair as alt
    alt.data_transformers.disable_max_rows()
except Exception:
    alt = None

_sortables_err = None
try:
    from streamlit_sortables import sort_items
except Exception as _se:
    sort_items = None
    _sortables_err = str(_se)

sys.path.insert(0, str(Path(__file__).parent))

OLLAMA_URL       = "http://localhost:11434"
DEFAULT_MODEL    = "qwen2.5:3b"
OVERRIDES_PATH   = Path("category_overrides.json")
TARGET_CATS_PATH = Path("target_categories.json")
CACHE_PATH       = Path("category_cache.json")
NAME_MATCHES_PATH = Path("name_matches.json")
CONFIG_PATH      = Path("config.json")
PRESETS_PATH     = Path("analytics_presets.json")

_pipe_err = None
try:
    import category_matcher as cm
    from pipeline import (
        read_table, get_col_map, SUPPORTED_EXTS,
        collect_source_files, read_and_anonymize, merge_anonymized,
        archive_source_files, load_config, save_config,
    )
except Exception as _e:
    import traceback
    _pipe_err = f"{type(_e).__name__}: {_e}\n\n{traceback.format_exc()}"


# ── shared helpers ───────────────────────────────────────────────────────────

def _ollama_ok(model):
    try:
        with urllib.request.urlopen(f"{OLLAMA_URL}/api/tags", timeout=2) as r:
            tags = json.loads(r.read())
        base = model.split(":")[0]
        return any(base in m["name"] for m in tags.get("models", []))
    except Exception:
        return False


def _folders():
    return sorted(d.name for d in Path(".").iterdir()
                  if d.is_dir() and not d.name.startswith(".")
                  and d.name not in ("__pycache__", ".venv"))


_NEW_FOLDER = "Создать новый каталог…"

def _folder_select(label, current, key):
    """Selectbox over existing folders + опция создания новой. Всегда mkdir выбранную."""
    opts = _folders()
    if current and current not in opts:
        opts = [current] + opts
    opts = opts + [_NEW_FOLDER]
    idx = opts.index(current) if current in opts else 0
    chosen = st.selectbox(label, opts, index=idx, key=key)
    if chosen == _NEW_FOLDER:
        typed = st.text_input(f"Имя новой папки ({label})", key=f"{key}_new",
                              placeholder="например: categories").strip()
        chosen = typed or current or "new_folder"
    if chosen and chosen != _NEW_FOLDER:
        try:
            Path(chosen).mkdir(parents=True, exist_ok=True)
        except Exception:
            pass
    return chosen


def _xlsx_bytes(df: pd.DataFrame) -> bytes:
    """Formatted xlsx (header style, freeze, auto width) into memory."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False)
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb.active
    if ws is not None:
        fill = PatternFill("solid", start_color="1F4E79")
        for cell in ws[1]:
            cell.font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 38
        ws.freeze_panes = "A2"
        for ci in range(1, ws.max_column + 1):
            letter = get_column_letter(ci)
            maxlen = max(
                (len(str(ws.cell(row=r, column=ci).value or ""))
                 for r in range(1, min(ws.max_row + 1, 200))),
                default=8,
            )
            ws.column_dimensions[letter].width = min(maxlen + 3, 52)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _load_corpus(folder):
    """[{name, brand, source_category, src_cats}] по уникальным (name, brand)."""
    seen = {}
    p = Path(folder)
    if not p.exists():
        return []
    for f in sorted(p.iterdir()):
        if f.suffix.lower() not in SUPPORTED_EXTS:
            continue
        try:
            df = read_table(f)
            df.columns = [str(c).strip() for c in df.columns]
            col = get_col_map(df)
            ic = col.get("item_name")
            if not ic or ic not in df.columns:
                continue
            bc    = col.get("brand")
            ccols = [c for c in (col.get(f"category_{k}") for k in range(5))
                     if c and c in df.columns]
            sub = df[[c for c in dict.fromkeys([ic, bc, *ccols]) if c]]
            for _, r in sub.iterrows():
                name = str(r[ic]).strip()
                if not name or name == "nan":
                    continue
                brand = str(r[bc]).strip() if bc else ""
                srcs  = [s for c in ccols if (s := str(r[c]).strip()) and s != "nan"]
                seen[(name, brand)] = {"name": name, "brand": brand,
                                       "source_category": " / ".join(srcs),
                                       "src_cats": srcs}
        except Exception as e:
            st.warning(f"{f.name}: {e}")
    return list(seen.values())


def _save_json(path, data):
    Path(path).write_text(json.dumps(data, ensure_ascii=False, indent=2), "utf-8")


def _read_json(path, default):
    p = Path(path)
    if not p.exists():
        return default
    try:
        return json.loads(p.read_text("utf-8"))
    except Exception:
        return default


# ══════════════════════════════════════════════════════════════════════════════
# TAB: ЗАПУСК
# ══════════════════════════════════════════════════════════════════════════════

def tab_run(cfg):
    st.markdown("#### Консолидация данных")
    st.caption("Чтение исходных таблиц, применение выбранных этапов обработки, "
               "объединение в единый набор данных и выгрузка результата.")

    in_dir = Path(cfg["input_folder"])
    files = collect_source_files(in_dir) if in_dir.exists() else []

    st.markdown("**Этапы обработки**")
    tg = st.columns(2)
    cfg["categorize"] = tg[0].toggle(
        "Категоризация", value=cfg["categorize"], key="run_categorize",
        help="Присвоение целевой категории каждому товару")
    cfg["anonymize"] = tg[1].toggle(
        "Обезличивание", value=cfg["anonymize"], key="run_anonymize",
        help="Маскировка магазинов, брендов, компаний и товаров")
    save_config(cfg)

    c1, c2, c3 = st.columns(3)
    c1.metric("Исходных файлов", len(files))
    c2.metric("Категоризация", "Включена" if cfg["categorize"] else "Отключена")
    c3.metric("Обезличивание", "Включено" if cfg["anonymize"] else "Отключено")

    if not files:
        st.warning(f"В каталоге «{cfg['input_folder']}» нет поддерживаемых файлов. "
                   "Загрузите их на вкладке «Файлы» или измените каталог в настройках.")
        return

    with st.expander(f"Файлы к обработке ({len(files)})"):
        for f in files:
            kb = max(f.stat().st_size // 1024, 1)
            st.markdown(f"- `{f.name}` · {kb} КБ")

    cols = st.columns([1, 1, 2])
    run_full   = cols[0].button("Запустить обработку", type="primary", use_container_width=True)
    archive_on = cols[1].checkbox("Архивировать исходники", value=False,
                                  help="Переместить обработанные файлы в каталог архива")

    if run_full:
        if cfg["name_match"] and not _ollama_ok(cfg["model"]):
            st.error("Сопоставление наименований включено, но сервис Ollama недоступен. "
                     "Отключите этап в настройках или запустите сервис.")
            return

        logbox = st.empty()
        logs = []
        def log(msg):
            logs.append(str(msg))
            logbox.code("\n".join(logs[-18:]))

        with st.status("Выполняется обработка…", expanded=True) as status:
            try:
                frames = read_and_anonymize(
                    files, log_fn=log,
                    anonymize=cfg["anonymize"],
                    categorize=cfg["categorize"],
                    name_match=cfg["name_match"],
                    model=cfg["model"],
                )
                if not frames:
                    status.update(label="Не удалось прочитать ни одного файла", state="error")
                    return
                merged = merge_anonymized(frames, log_fn=log,
                                          aggregate=cfg["aggregate_after_merge"])
                st.session_state.merged_df = merged
                st.session_state.merged_frames_info = [(n, len(d)) for n, d in frames]

                if archive_on:
                    archive_source_files(files, cfg["archive_folder"], log_fn=log)

                status.update(label=f"Обработка завершена: {len(merged):,} строк",
                              state="complete")
            except Exception as e:
                import traceback
                log(traceback.format_exc())
                status.update(label=f"Ошибка обработки: {e}", state="error")
                return

    # ── результат ──
    merged = st.session_state.get("merged_df")
    if merged is None:
        return

    st.divider()
    st.markdown("#### Результат")
    m1, m2, m3 = st.columns(3)
    m1.metric("Строк", f"{len(merged):,}")
    m2.metric("Столбцов", len(merged.columns))
    if "target_category" in merged.columns:
        filled = (merged["target_category"].astype(str).str.strip() != "").sum()
        m3.metric("Покрытие категориями", f"{filled / len(merged):.0%}")

    st.dataframe(merged.head(200), use_container_width=True, height=320)

    default_name = f"merged_{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
    cda, cdb = st.columns([2, 1])
    fname = cda.text_input("Имя файла", default_name, key="dl_name")
    if not fname.endswith(".xlsx"):
        fname += ".xlsx"

    with st.spinner("Формирование файла…"):
        xbytes = _xlsx_bytes(merged)

    cdb.download_button(
        "Скачать XLSX", data=xbytes, file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary", use_container_width=True,
    )

    csv = merged.to_csv(index=False).encode("utf-8-sig")
    st.download_button("Скачать CSV", data=csv,
                       file_name=fname.replace(".xlsx", ".csv"), mime="text/csv")

    if st.button("Сохранить в каталог результатов"):
        out_dir = Path(cfg["output_folder"]); out_dir.mkdir(parents=True, exist_ok=True)
        (out_dir / fname).write_bytes(xbytes)
        st.success(f"Файл сохранён: {(out_dir / fname).resolve()}")


# ══════════════════════════════════════════════════════════════════════════════
# TAB: КАТЕГОРИИ
# ══════════════════════════════════════════════════════════════════════════════

_CAT_DF_COLS = ["name", "brand", "source_category", "target_category", "confidence"]


def _run_categorize(rows, cats, prog):
    cache, ov = cm.load_cache(), cm.load_overrides()
    todo, keys = [], []
    for r in rows:
        k = cm._key(r["name"], r["brand"])
        if k in cache or k in ov:
            continue
        todo.append({"name": r["name"], "brand": r["brand"], "src_cats": r["src_cats"]})
        keys.append(k)
    if not todo:
        prog("все товары уже обработаны")
        return
    res = cm.classify_batch(todo, cats)
    for k, rr in zip(keys, res):
        cache[k] = {"target_category": rr["target_category"],
                    "category_confidence": rr["category_confidence"]}
    cm.save_cache(cache)
    prog(f"обработано новых товаров: {len(todo)}")


def _build_cat_df(rows):
    cache, ov = cm.load_cache(), cm.load_overrides()
    recs = []
    for r in rows:
        k = cm._key(r["name"], r["brand"])
        if k in ov:
            tc, conf = ov[k].get("target_category", ""), float(ov[k].get("category_confidence", 1.0))
        else:
            c = cache.get(k, {})
            tc, conf = c.get("target_category", ""), c.get("category_confidence", 0.0)
        recs.append({"name": r["name"], "brand": r["brand"],
                     "source_category": r["source_category"],
                     "target_category": tc, "confidence": conf})
    if not recs:
        return pd.DataFrame(columns=_CAT_DF_COLS)
    return pd.DataFrame(recs)


def tab_categorize(cfg):
    cats = cm.load_target_categories()

    with st.expander(f"Целевые категории ({len(cats)})", expanded=not cats):
        st.caption("Перечень целевых категорий, по одной в строке. "
                   "Сопоставление товаров выполняется автоматически.")
        edited = st.text_area("Категории", "\n".join(cats), height=220,
                              label_visibility="collapsed")
        cc1, cc2 = st.columns(2)
        if cc1.button("Сохранить категории", use_container_width=True):
            new = [c.strip() for c in edited.splitlines() if c.strip()]
            _save_json(TARGET_CATS_PATH, new)
            st.toast(f"Сохранено категорий: {len(new)}"); st.rerun()
        if cc2.button("Импортировать из источника", use_container_width=True,
                      help="Извлекает уникальные значения верхнего уровня категорий "
                           "из файлов в каталоге «Источник категорий»"):
            found = _suggest_categories(cfg["category_folder"])
            if found:
                _save_json(TARGET_CATS_PATH, found)
                st.toast(f"Импортировано категорий: {len(found)}"); st.rerun()
            else:
                st.warning("В источнике не найдено столбцов с категориями.")

    if not cats:
        st.info("Задайте целевые категории выше или импортируйте их из источника.")
        return

    c1, c2 = st.columns([3, 1])
    with c1:
        folder = _folder_select("Каталог для классификации", cfg["category_folder"], "cat_corpus")
    if c2.button("Классифицировать", type="primary", use_container_width=True):
        rows = _load_corpus(folder)
        prog = st.empty()
        _run_categorize(rows, cats, lambda s: prog.caption(f"Классификация: {s}"))
        prog.empty()
        st.session_state.cat_df = _build_cat_df(rows)
        st.toast(f"Обработано товаров: {len(rows)}")

    if "cat_df" not in st.session_state:
        st.caption("Нажмите «Классифицировать» для обработки каталога.")
        return

    df = st.session_state.cat_df
    if df.empty or "confidence" not in df.columns:
        st.info("Каталог пуст или в нём отсутствует столбец с наименованием товара.")
        st.session_state.pop("cat_df", None)
        return

    f1, f2 = st.columns([1, 1])
    only_low   = f1.checkbox("Только с уверенностью ниже 70%", key="cat_low")
    only_empty = f2.checkbox("Только без категории", key="cat_empty")
    view = df
    if only_low:
        view = view[view["confidence"] < 0.70]
    if only_empty:
        view = view[view["target_category"].astype(str).str.strip() == ""]
    st.caption(f"Отображено {len(view)} из {len(df)} · "
               f"без категории: {(df['target_category'].astype(str).str.strip() == '').sum()} · "
               f"низкая уверенность: {(df['confidence'] < 0.70).sum()}")

    edited = st.data_editor(
        view, hide_index=True, use_container_width=True, height=460,
        key="cat_editor",
        column_config={
            "name":            st.column_config.TextColumn("Наименование", disabled=True, width="large"),
            "brand":           st.column_config.TextColumn("Бренд", disabled=True),
            "source_category": st.column_config.TextColumn("Категория источника", disabled=True),
            "target_category": st.column_config.SelectboxColumn(
                "Целевая категория", options=[""] + cats, required=False),
            "confidence":      st.column_config.NumberColumn("Уверенность", disabled=True, format="%.0f%%"),
        },
    )

    b1, b2 = st.columns([1, 1])
    if b1.button("Применить изменения", type="primary", key="cat_apply", use_container_width=True):
        orig = view.set_index(["name", "brand"])["target_category"]
        new  = edited.set_index(["name", "brand"])["target_category"]
        ov = cm.load_overrides()
        changed = 0
        for key in new.index:
            if str(new[key]) != str(orig.get(key, "")):
                ov[cm._key(*key)] = {"target_category": new[key], "category_confidence": 1.0}
                changed += 1
        if changed:
            _save_json(OVERRIDES_PATH, ov)
            st.session_state.pop("cat_df", None)
            st.toast(f"Сохранено изменений: {changed}"); st.rerun()
        else:
            st.toast("Изменений не обнаружено")

    cat_csv = df.to_csv(index=False).encode("utf-8-sig")
    b2.download_button("Скачать таблицу категорий (CSV)", data=cat_csv,
                       file_name="categories.csv", mime="text/csv",
                       use_container_width=True)


def _suggest_categories(folder):
    """Уникальные значения самого верхнего заполненного уровня категорий."""
    p = Path(folder)
    if not p.exists():
        return []
    found = set()
    for f in sorted(p.iterdir()):
        if f.suffix.lower() not in SUPPORTED_EXTS:
            continue
        try:
            df = read_table(f)
            df.columns = [str(c).strip() for c in df.columns]
            col = get_col_map(df)
            for k in range(5):
                c = col.get(f"category_{k}")
                if c and c in df.columns:
                    vals = df[c].dropna().astype(str).str.strip()
                    vals = vals[(vals != "") & (vals != "nan")]
                    if len(vals):
                        found.update(vals.unique())
                        break
        except Exception:
            continue
    return sorted(found)


# ══════════════════════════════════════════════════════════════════════════════
# TAB: АНАЛИТИКА
# ══════════════════════════════════════════════════════════════════════════════

def _num(s):
    return pd.to_numeric(s, errors="coerce").fillna(0)


# палитра для altair-графиков
_SCHEME = "tableau20"
_BAR_COLOR = "#4C78D9"


def _analytics_source(cfg):
    """Возвращает DataFrame для аналитики (сессия / файл / загрузка) или None."""
    src = st.radio("Источник данных",
                   ["Из последнего запуска", "Из файла результатов", "Загрузить файл"],
                   horizontal=True, key="an_src")
    if src == "Из последнего запуска":
        df = st.session_state.get("merged_df")
        if df is None:
            st.info("Сначала выполните обработку на вкладке «Обработка».")
        return df
    if src == "Из файла результатов":
        out = Path(cfg["output_folder"])
        files = sorted([f for f in out.iterdir() if f.suffix.lower() in SUPPORTED_EXTS],
                       reverse=True) if out.exists() else []
        if not files:
            st.info(f"В «{cfg['output_folder']}/» нет таблиц.")
            return None
        pick = st.selectbox("Файл", [f.name for f in files], key="an_file")
        df = read_table(out / pick)
        df.columns = [str(c).strip() for c in df.columns]
        return df
    up = st.file_uploader("xlsx/csv", type=["xlsx", "csv", "tsv"], key="an_up")
    return read_table_buffer(up) if up else None


def _period_series(df):
    """'YYYY-MM' (сортируемый ключ) для оси времени. Векторизовано."""
    if "month" not in df.columns:
        return None
    mo = _num(df["month"]).astype(int)
    if "year" in df.columns:
        yr = _num(df["year"]).astype(int)
        return yr.astype(str) + "-" + mo.map("{:02d}".format)
    return mo.map("{:02d}".format)


def _render_abc_xyz(fdf, name_col, short_label, unit):
    if not name_col:
        st.info("В данных нет столбца с наименованием товара."); return
    g = fdf.groupby(name_col)["_val"].sum().sort_values(ascending=False)
    g = g[g > 0]
    if g.empty:
        st.info("Недостаточно данных для анализа."); return
    cum = g.cumsum() / g.sum()
    abc = pd.Series(np.where(cum <= 0.8, "A", np.where(cum <= 0.95, "B", "C")), index=g.index)
    res = pd.DataFrame({short_label: g, "ABC": abc})

    period = _period_series(fdf)
    xyz = None
    if period is not None and period.nunique() >= 3:
        pv = (fdf.assign(_p=period)
                 .pivot_table(index=name_col, columns="_p", values="_val", aggfunc="sum")
                 .fillna(0))
        mean = pv.mean(axis=1)
        cv = (pv.std(axis=1) / mean.replace(0, np.nan)).fillna(np.inf)
        xyz = pd.Series(np.where(cv <= 0.1, "X", np.where(cv <= 0.25, "Y", "Z")), index=pv.index)
        res["XYZ"] = res.index.map(xyz).fillna("Z")

    st.markdown("**Классы ABC** — A: 80% оборота, B: 15%, C: 5%")
    summary = res.groupby("ABC")[short_label].agg(Товаров="count", Сумма="sum")
    summary["Доля, %"] = (summary["Сумма"] / summary["Сумма"].sum() * 100).round(1)
    cc = st.columns([1, 1.3])
    cc[0].dataframe(summary.style.format({"Сумма": "{:,.0f}", "Доля, %": "{:.1f}"}),
                    use_container_width=True)
    if alt is not None:
        sg = summary.reset_index()
        ch = (alt.Chart(sg).mark_arc(innerRadius=55)
              .encode(theta="Сумма:Q",
                      color=alt.Color("ABC:N", scale=alt.Scale(domain=["A", "B", "C"],
                                      range=["#16A34A", "#F59E0B", "#DC2626"])),
                      tooltip=["ABC:N", alt.Tooltip("Сумма:Q", format=",.0f"), "Товаров:Q"])
              .properties(height=260))
        cc[1].altair_chart(ch, use_container_width=True)

    if xyz is not None:
        st.markdown("**Матрица ABC × XYZ** — XYZ отражает стабильность спроса "
                    "(X — стабильный, Y — колеблющийся, Z — нерегулярный)")
        mtx = res.pivot_table(index="ABC", columns="XYZ", values=short_label,
                              aggfunc="count", fill_value=0)
        st.dataframe(mtx, use_container_width=True)
    else:
        st.caption("XYZ-анализ требует не менее 3 периодов — для текущего набора недоступен.")

    st.dataframe(res.head(300).style.format({short_label: "{:,.0f}"}),
                 use_container_width=True, height=300)
    st.download_button("Скачать результат ABC/XYZ (CSV)",
                       res.to_csv().encode("utf-8-sig"), "abc_xyz.csv", "text/csv",
                       key="dlf_abcxyz")


def _render_compare(fdf, cat_col, short_label):
    period = _period_series(fdf)
    if period is None or period.nunique() < 2:
        st.info("Для сравнения требуется не менее 2 периодов в данных."); return
    fdf = fdf.assign(_p=period)
    periods = sorted(fdf["_p"].unique())
    c = st.columns(2)
    p1 = c[0].selectbox("Период A", periods, index=0, key="cmp_a")
    p2 = c[1].selectbox("Период B", periods, index=len(periods) - 1, key="cmp_b")
    dim = cat_col if cat_col else ("brand" if "brand" in fdf.columns else None)
    if dim is None:
        st.info("В данных нет измерения для сравнения (категория или бренд)."); return

    a = fdf[fdf["_p"] == p1].groupby(dim)["_val"].sum()
    b = fdf[fdf["_p"] == p2].groupby(dim)["_val"].sum()
    comp = pd.DataFrame({p1: a, p2: b}).fillna(0)
    comp["Δ"] = comp[p2] - comp[p1]
    comp["Δ, %"] = np.where(comp[p1] != 0, comp["Δ"] / comp[p1] * 100, np.nan)
    comp = comp.sort_values("Δ", ascending=False)
    st.dataframe(comp.style.format({p1: "{:,.0f}", p2: "{:,.0f}",
                                    "Δ": "{:+,.0f}", "Δ, %": "{:+.1f}"}),
                 use_container_width=True, height=340)
    if alt is not None:
        mv = comp.reset_index().rename(columns={dim: "dim"})
        top = pd.concat([mv.nlargest(10, "Δ"), mv.nsmallest(10, "Δ")]).drop_duplicates("dim")
        ch = (alt.Chart(top).mark_bar()
              .encode(x=alt.X("Δ:Q", title="Изменение", axis=alt.Axis(format="~s")),
                      y=alt.Y("dim:N", sort="-x", title=None),
                      color=alt.condition("datum['Δ'] >= 0", alt.value("#16A34A"), alt.value("#DC2626")),
                      tooltip=["dim:N", alt.Tooltip("Δ:Q", format="+,.0f")])
              .properties(height=max(26 * len(top) + 30, 200)))
        st.altair_chart(ch, use_container_width=True)
    st.download_button("Скачать сравнение (CSV)", comp.to_csv().encode("utf-8-sig"),
                       "comparison.csv", "text/csv", key="dlf_compare")


def _render_anomalies(fdf, name_col, short_label):
    found = False
    if {"sales_rub", "cost_rub"}.issubset(fdf.columns) and name_col:
        g = fdf.groupby(name_col).agg(
            Выручка=("sales_rub", lambda s: _num(s).sum()),
            Себестоимость=("cost_rub", lambda s: _num(s).sum()))
        g["Маржа, %"] = np.where(g["Выручка"] != 0,
                                 (g["Выручка"] - g["Себестоимость"]) / g["Выручка"] * 100, np.nan)
        low = g[g["Маржа, %"] < 5].sort_values("Маржа, %")
        st.markdown("**Низкая или отрицательная маржа** (ниже 5%)")
        if len(low):
            st.dataframe(low.head(100).style.format({"Выручка": "{:,.0f}",
                         "Себестоимость": "{:,.0f}", "Маржа, %": "{:.1f}"}),
                         use_container_width=True, height=240)
            found = True
        else:
            st.caption("Не обнаружено.")

    period = _period_series(fdf)
    if period is not None and period.nunique() >= 3 and name_col:
        pv = (fdf.assign(_p=period)
                 .pivot_table(index=name_col, columns="_p", values="_val", aggfunc="sum")
                 .fillna(0))
        last, base = pv.iloc[:, -1], pv.iloc[:, :-1]
        mean, std = base.mean(axis=1), base.std(axis=1).replace(0, np.nan)
        z = (last - mean) / std
        anom = pd.DataFrame({"Текущий период": last, "Среднее": mean, "Z-оценка": z}).dropna()
        anom = anom[anom["Z-оценка"].abs() >= 2].sort_values("Z-оценка")
        st.markdown("**Аномальные отклонения в последнем периоде** (|Z| ≥ 2)")
        if len(anom):
            st.dataframe(anom.style.format({"Текущий период": "{:,.0f}", "Среднее": "{:,.0f}",
                         "Z-оценка": "{:+.2f}"}), use_container_width=True, height=240)
            found = True
        else:
            st.caption("Не обнаружено.")
    else:
        st.caption("Анализ всплесков и провалов требует не менее 3 периодов.")

    if not found:
        st.info("Аномалии не выявлены или данных недостаточно.")


def _render_forecast(fdf, short_label, unit):
    period = _period_series(fdf)
    if period is None or period.nunique() < 2:
        st.info("Прогноз требует не менее 2 периодов в данных."); return
    g = fdf.assign(_p=period).groupby("_p")["_val"].sum().sort_index()
    periods, y = list(g.index), g.values.astype(float)
    x = np.arange(len(y))
    coef = np.polyfit(x, y, 1)
    horizon = st.slider("Горизонт прогноза (периодов)", 1, 6, 3, key="fc_h")
    fy = np.clip(np.polyval(coef, np.arange(len(y), len(y) + horizon)), 0, None)
    flabels = [f"прогноз +{i + 1}" for i in range(horizon)]
    alld = pd.concat([
        pd.DataFrame({"Период": periods, "Значение": y, "Тип": "факт"}),
        pd.DataFrame({"Период": flabels, "Значение": fy, "Тип": "прогноз"}),
    ])
    if alt is not None:
        ch = (alt.Chart(alld).mark_line(point=True)
              .encode(x=alt.X("Период:N", sort=list(alld["Период"])),
                      y=alt.Y("Значение:Q", axis=alt.Axis(format="~s")),
                      color=alt.Color("Тип:N", scale=alt.Scale(domain=["факт", "прогноз"],
                                      range=[_BAR_COLOR, "#F59E0B"])),
                      strokeDash=alt.StrokeDash("Тип:N", legend=None),
                      tooltip=["Период:N", "Тип:N", alt.Tooltip("Значение:Q", format=",.0f")])
              .properties(height=320))
        st.altair_chart(ch, use_container_width=True)
    trend = "рост" if coef[0] > 0 else "снижение" if coef[0] < 0 else "без изменений"
    st.caption(f"Линейный тренд: {trend} ≈ {coef[0]:+,.0f} {unit} за период. "
               f"Суммарный прогноз на {horizon} периодов: {fy.sum():,.0f} {unit}.")
    st.caption("Прогноз основан на линейной экстраполяции; точность растёт с числом периодов.")


def _excel_report_bytes(fdf, cat_col, name_col, short_label):
    buf = io.BytesIO()
    metric_cols = {}
    if "sales_rub" in fdf.columns: metric_cols["Выручка"] = _num(fdf["sales_rub"])
    if "qty_sold" in fdf.columns:  metric_cols["Количество"] = _num(fdf["qty_sold"])
    if "cost_rub" in fdf.columns:  metric_cols["Себестоимость"] = _num(fdf["cost_rub"])

    def _by(dim):
        t = fdf[[dim]].copy()
        for n, s in metric_cols.items():
            t[n] = s.values
        out = t.groupby(dim).sum()
        if "Выручка" in out and "Себестоимость" in out:
            out["Прибыль"] = out["Выручка"] - out["Себестоимость"]
        return out.sort_values(out.columns[0], ascending=False)

    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        kpi = pd.DataFrame({
            "Показатель": ["Строк", f"Итого {short_label}", "Уникальных товаров"],
            "Значение": [len(fdf), fdf["_val"].sum(),
                         fdf[name_col].nunique() if name_col else "—"],
        })
        kpi.to_excel(w, sheet_name="Сводка", index=False)
        if cat_col: _by(cat_col).to_excel(w, sheet_name="Категории")
        if "brand" in fdf.columns: _by("brand").head(100).to_excel(w, sheet_name="Бренды")
        if "region" in fdf.columns: _by("region").to_excel(w, sheet_name="Регионы")
        period = _period_series(fdf)
        if period is not None:
            fdf.assign(_p=period).groupby("_p")["_val"].sum().to_frame(short_label) \
                .to_excel(w, sheet_name="Динамика")
        if name_col:
            fdf.groupby(name_col)["_val"].sum().sort_values(ascending=False).head(200) \
                .to_frame(short_label).to_excel(w, sheet_name="Топ товаров")
    return buf.getvalue()


def _pdf_report_bytes(fdf, cat_col, name_col, short_label, unit):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages

    buf = io.BytesIO()
    with PdfPages(buf) as pdf:
        fig = plt.figure(figsize=(8.27, 11.69))
        fig.text(0.5, 0.93, "EnigmaSQL — аналитический отчёт", ha="center",
                 fontsize=18, weight="bold")
        lines = [f"Показатель: {short_label}",
                 f"Итого: {fdf['_val'].sum():,.0f} {unit}",
                 f"Строк данных: {len(fdf):,}"]
        if name_col:
            lines.append(f"Уникальных товаров: {fdf[name_col].nunique():,}")
        if "store_name" in fdf.columns:
            lines.append(f"Магазинов: {fdf['store_name'].nunique():,}")
        fig.text(0.12, 0.82, "\n".join(lines), fontsize=12, va="top")
        plt.axis("off"); pdf.savefig(fig); plt.close(fig)

        def _barpage(series, title, color):
            series = series[series.index.astype(str).str.strip() != ""]
            if series.empty:
                return
            fig, ax = plt.subplots(figsize=(8.27, 5.2))
            series[::-1].plot.barh(ax=ax, color=color)
            ax.set_title(title); fig.tight_layout()
            pdf.savefig(fig); plt.close(fig)

        if cat_col:
            _barpage(fdf.groupby(cat_col)["_val"].sum().sort_values(ascending=False).head(15),
                     f"{short_label} по категориям", "#4C78D9")
        if "brand" in fdf.columns:
            _barpage(fdf.groupby("brand")["_val"].sum().sort_values(ascending=False).head(15),
                     "Ведущие бренды", "#8B5CF6")
        period = _period_series(fdf)
        if period is not None and period.nunique() > 1:
            g = fdf.assign(_p=period).groupby("_p")["_val"].sum().sort_index()
            fig, ax = plt.subplots(figsize=(8.27, 4.2))
            g.plot(marker="o", ax=ax, color="#22C55E")
            ax.set_title("Динамика по периодам"); fig.tight_layout()
            pdf.savefig(fig); plt.close(fig)
    return buf.getvalue()


def tab_analytics(cfg):
    st.markdown("#### Аналитика продаж")

    df = _analytics_source(cfg)
    if df is None or df.empty:
        if df is not None:
            st.warning("Таблица пуста.")
        return

    df = df.copy()
    has = lambda c: c in df.columns
    name_col = ("canonical_name" if has("canonical_name")
                else "item_name" if has("item_name") else None)
    cat_col = "target_category" if has("target_category") else None

    # ── метрики (выбор показателя) ──
    metric_opts = {}
    if has("sales_rub"): metric_opts["Выручка, ₽"] = "sales_rub"
    if has("qty_sold"):  metric_opts["Количество, шт"] = "qty_sold"
    if has("cost_rub"):  metric_opts["Себестоимость, ₽"] = "cost_rub"
    if has("sales_rub") and has("cost_rub"): metric_opts["Прибыль, ₽"] = "__profit"
    if not metric_opts:
        st.warning("В таблице нет числовых метрик (sales_rub / qty_sold).")
        st.dataframe(df.head(200), use_container_width=True)
        return

    # ── ФИЛЬТРЫ ──
    filt_keys = {"an_metric", "an_f_year", "an_f_region", f"an_f_{cat_col}",
                 "an_f2_brand", "an_f2_store_name", "an_f2_month", "an_topn"}
    with st.expander("Фильтры", expanded=True):
        # пресеты
        presets = _read_json(PRESETS_PATH, {})
        pc = st.columns([2, 1, 1])
        chosen = pc[0].selectbox("Сохранённый набор фильтров",
                                 ["—"] + sorted(presets), key="an_preset_pick")
        if pc[1].button("Применить набор", use_container_width=True, key="mv_preset_apply"):
            if chosen != "—" and chosen in presets:
                for kk, vv in presets[chosen].items():
                    st.session_state[kk] = vv
                st.rerun()
        if pc[2].button("Сбросить фильтры", use_container_width=True, key="warn_preset_reset"):
            for kk in list(filt_keys):
                st.session_state.pop(kk, None)
            st.rerun()
        pname = st.text_input("Имя нового набора", key="an_preset_name",
                              placeholder="например: Москва · стирка · 2026")
        if st.button("Сохранить текущие фильтры", key="mv_preset_save"):
            if pname.strip():
                presets[pname.strip()] = {kk: st.session_state.get(kk)
                                          for kk in filt_keys if kk in st.session_state}
                _save_json(PRESETS_PATH, presets)
                st.toast(f"Набор сохранён: {pname.strip()}"); st.rerun()

        st.divider()
        fc = st.columns(4)
        metric_label = fc[0].selectbox("Показатель", list(metric_opts), key="an_metric")
        metric = metric_opts[metric_label]

        def _ms(container, label, column, key):
            if not has(column):
                return None
            opts = sorted(df[column].dropna().astype(str).str.strip().unique())
            opts = [o for o in opts if o and o != "nan"]
            if not opts:
                return None
            return container.multiselect(label, opts, key=key)

        sel_year   = _ms(fc[1], "Год", "year", "an_f_year")
        sel_region = _ms(fc[2], "Регион", "region", "an_f_region")
        sel_cat    = _ms(fc[3], "Категория", cat_col, f"an_f_{cat_col}") if cat_col else None
        fc2 = st.columns(4)
        sel_brand = _ms(fc2[0], "Бренд", "brand", "an_f2_brand")
        sel_store = _ms(fc2[1], "Магазин", "store_name", "an_f2_store_name")
        sel_month = _ms(fc2[2], "Месяц", "month", "an_f2_month")
        topn = fc2[3].slider("Топ-N в рейтингах", 5, 30, 15, key="an_topn")

    # применяем фильтры
    mask = pd.Series(True, index=df.index)
    def _apply(column, sel):
        nonlocal mask
        if sel:
            mask &= df[column].astype(str).str.strip().isin(sel)
    _apply("year", sel_year); _apply("region", sel_region)
    if cat_col: _apply(cat_col, sel_cat)
    _apply("brand", sel_brand); _apply("store_name", sel_store)
    _apply("month", sel_month)
    fdf = df[mask].copy()

    if fdf.empty:
        st.warning("Под заданные фильтры данные отсутствуют — измените условия.")
        return

    # значение метрики
    if metric == "__profit":
        fdf["_val"] = _num(fdf["sales_rub"]) - _num(fdf["cost_rub"])
    else:
        fdf["_val"] = _num(fdf[metric])
    unit = "шт" if metric == "qty_sold" else "₽"
    short_label = metric_label.split(",")[0].strip()

    st.caption(f"Отобрано **{len(fdf):,}** из {len(df):,} строк · показатель: **{metric_label}**")

    # ── KPI ──
    total = fdf["_val"].sum()
    k = st.columns(5)
    k[0].metric(f"{short_label}, итого", f"{total:,.0f} {unit}")
    if name_col:
        k[1].metric("Уникальных товаров", f"{fdf[name_col].nunique():,}")
        avg = total / max(fdf[name_col].nunique(), 1)
        k[2].metric("В среднем на товар", f"{avg:,.0f} {unit}")
    if has("store_name"):
        k[3].metric("Магазинов", f"{fdf['store_name'].nunique():,}")
    if has("sales_rub") and has("cost_rub"):
        rev, cost = _num(fdf["sales_rub"]).sum(), _num(fdf["cost_rub"]).sum()
        k[4].metric("Валовая маржа", f"{(rev - cost) / rev * 100:.1f}%" if rev else "—")

    st.divider()

    # ════════ СТРУКТУРА: доли категорий + распределение ════════
    if cat_col and alt is not None:
        st.markdown(f"##### Структура продаж по категориям")
        g = (fdf.groupby(cat_col)["_val"].sum().reset_index()
                .rename(columns={cat_col: "Категория", "_val": "Значение"}))
        g = g[g["Категория"].astype(str).str.strip() != ""]
        if not g.empty:
            g["Доля"] = g["Значение"] / g["Значение"].sum()
            cdo = st.columns([1, 1.3])
            donut = (alt.Chart(g).mark_arc(innerRadius=70, stroke="#fff", strokeWidth=2)
                     .encode(
                         theta=alt.Theta("Значение:Q", stack=True),
                         color=alt.Color("Категория:N",
                                         scale=alt.Scale(scheme=_SCHEME),
                                         legend=alt.Legend(orient="right", labelLimit=180)),
                         tooltip=["Категория:N", alt.Tooltip("Значение:Q", format=",.0f"),
                                  alt.Tooltip("Доля:Q", format=".1%")])
                     .properties(height=340))
            cdo[0].altair_chart(donut, use_container_width=True)

            bar = (alt.Chart(g.sort_values("Значение", ascending=False))
                   .mark_bar(cornerRadiusEnd=4, color=_BAR_COLOR)
                   .encode(
                       x=alt.X("Значение:Q", title=short_label, axis=alt.Axis(format="~s")),
                       y=alt.Y("Категория:N", sort="-x", title=None),
                       tooltip=["Категория:N", alt.Tooltip("Значение:Q", format=",.0f"),
                                alt.Tooltip("Доля:Q", format=".1%")])
                   .properties(height=340))
            cdo[1].altair_chart(bar, use_container_width=True)
        empty = (fdf[cat_col].astype(str).str.strip() == "").mean()
        st.caption(f"Покрытие категориями: **{(1 - empty):.0%}** строк")
        st.divider()

    # ════════ ТОПЫ: бренды и товары ════════
    cc = st.columns(2)
    def _top_bar(container, column, title, n):
        if not (has(column) and alt is not None):
            return
        g = (fdf.groupby(column)["_val"].sum().sort_values(ascending=False)
                .head(n).reset_index())
        g.columns = ["label", "val"]
        g["label"] = g["label"].astype(str).str.slice(0, 42)
        ch = (alt.Chart(g).mark_bar(cornerRadiusEnd=4)
              .encode(
                  x=alt.X("val:Q", title=short_label, axis=alt.Axis(format="~s")),
                  y=alt.Y("label:N", sort="-x", title=None),
                  color=alt.Color("val:Q", scale=alt.Scale(scheme="blues"), legend=None),
                  tooltip=[alt.Tooltip("label:N", title=title),
                           alt.Tooltip("val:Q", format=",.0f", title=short_label)])
              .properties(height=max(28 * len(g) + 30, 200)))
        container.markdown(f"##### {title}")
        container.altair_chart(ch, use_container_width=True)

    _top_bar(cc[0], "brand", f"Ведущие бренды (топ-{topn})", topn)
    if name_col:
        _top_bar(cc[1], name_col, f"Ведущие товары (топ-{topn})", topn)

    # ════════ ПАРЕТО-АНАЛИЗ ════════
    if name_col and alt is not None:
        st.markdown("##### Парето-анализ: вклад товаров в оборот")
        g = fdf.groupby(name_col)["_val"].sum().sort_values(ascending=False).reset_index()
        g.columns = ["label", "val"]
        g = g.head(40)
        g["cum"] = g["val"].cumsum() / fdf["_val"].sum()
        g["rank"] = range(1, len(g) + 1)
        g["label"] = g["label"].astype(str).str.slice(0, 30)
        base = alt.Chart(g).encode(x=alt.X("label:N", sort=g["label"].tolist(),
                                           title=None, axis=alt.Axis(labelAngle=-45, labelLimit=120)))
        bars = base.mark_bar(color=_BAR_COLOR).encode(
            y=alt.Y("val:Q", title=short_label, axis=alt.Axis(format="~s")),
            tooltip=["label:N", alt.Tooltip("val:Q", format=",.0f")])
        line = base.mark_line(color="#E45756", point=True).encode(
            y=alt.Y("cum:Q", title="Накопленная доля", axis=alt.Axis(format="%")),
            tooltip=[alt.Tooltip("cum:Q", format=".1%", title="Накопленная доля")])
        st.altair_chart(alt.layer(bars, line).resolve_scale(y="independent")
                        .properties(height=340), use_container_width=True)

    # ════════ ГЕОГРАФИЯ ════════
    if has("region") and alt is not None:
        st.divider()
        st.markdown(f"##### {short_label} по регионам")
        g = (fdf.groupby("region")["_val"].sum().sort_values(ascending=False).reset_index())
        g.columns = ["region", "val"]
        ch = (alt.Chart(g).mark_bar(cornerRadiusEnd=4)
              .encode(
                  x=alt.X("val:Q", title=short_label, axis=alt.Axis(format="~s")),
                  y=alt.Y("region:N", sort="-x", title=None),
                  color=alt.Color("val:Q", scale=alt.Scale(scheme="greens"), legend=None),
                  tooltip=["region:N", alt.Tooltip("val:Q", format=",.0f")])
              .properties(height=max(30 * len(g) + 30, 200)))
        st.altair_chart(ch, use_container_width=True)

        # тепловая карта категория × регион
        if cat_col:
            st.markdown("##### Распределение по категориям и регионам")
            g2 = (fdf.groupby([cat_col, "region"])["_val"].sum().reset_index())
            g2 = g2[g2[cat_col].astype(str).str.strip() != ""]
            heat = (alt.Chart(g2).mark_rect()
                    .encode(
                        x=alt.X("region:N", title=None),
                        y=alt.Y(f"{cat_col}:N", title=None),
                        color=alt.Color("_val:Q", scale=alt.Scale(scheme="yelloworangered"),
                                        legend=alt.Legend(title=short_label, format="~s")),
                        tooltip=[f"{cat_col}:N", "region:N",
                                 alt.Tooltip("_val:Q", format=",.0f")])
                    .properties(height=max(22 * g2[cat_col].nunique() + 40, 200)))
            st.altair_chart(heat, use_container_width=True)

    # ════════ ДИНАМИКА ════════
    period = _period_series(fdf)
    if period is not None and alt is not None and period.nunique() > 1:
        st.divider()
        st.markdown(f"##### Динамика по периодам")
        fdf["_period"] = period
        if cat_col:
            g = (fdf.groupby(["_period", cat_col])["_val"].sum().reset_index())
            g = g[g[cat_col].astype(str).str.strip() != ""]
            area = (alt.Chart(g).mark_area(opacity=0.85)
                    .encode(
                        x=alt.X("_period:N", title=None),
                        y=alt.Y("_val:Q", stack=True, title=short_label, axis=alt.Axis(format="~s")),
                        color=alt.Color(f"{cat_col}:N", scale=alt.Scale(scheme=_SCHEME),
                                        legend=alt.Legend(orient="bottom", labelLimit=140)),
                        tooltip=["_period:N", f"{cat_col}:N",
                                 alt.Tooltip("_val:Q", format=",.0f")])
                    .properties(height=340))
            st.altair_chart(area, use_container_width=True)
        else:
            g = fdf.groupby("_period")["_val"].sum().reset_index()
            line = (alt.Chart(g).mark_line(point=True, color=_BAR_COLOR)
                    .encode(x="_period:N", y=alt.Y("_val:Q", axis=alt.Axis(format="~s")),
                            tooltip=["_period:N", alt.Tooltip("_val:Q", format=",.0f")])
                    .properties(height=300))
            st.altair_chart(line, use_container_width=True)

    # ════════ МАГАЗИНЫ ════════
    if has("store_name"):
        st.divider()
        st.markdown("##### Рейтинг магазинов")
        col_total = f"{short_label}, итого"
        gr = fdf.groupby("store_name").agg(**{col_total: ("_val", "sum")})
        if name_col:
            gr["Товаров"] = fdf.groupby("store_name")[name_col].nunique()
        gr = gr.sort_values(col_total, ascending=False).head(50)
        gr[col_total] = gr[col_total].round(0)
        st.dataframe(gr.style.format({col_total: "{:,.0f}"}),
                     use_container_width=True, height=300)

    # ════════ СВОДНАЯ ТАБЛИЦА + ВЫГРУЗКА ════════
    st.divider()
    st.markdown("##### Сводная таблица и выгрузка")
    pivot = None
    if cat_col:
        cols = {}
        if has("sales_rub"): cols["Выручка"] = _num(fdf["sales_rub"])
        if has("qty_sold"):  cols["Количество"] = _num(fdf["qty_sold"])
        if has("cost_rub"):  cols["Себестоимость"] = _num(fdf["cost_rub"])
        if cols:
            tmp = fdf[[cat_col]].copy()
            for name, series in cols.items():
                tmp[name] = series
            pivot = tmp.groupby(cat_col).sum()
            if "Выручка" in pivot and "Себестоимость" in pivot:
                pivot["Прибыль"] = pivot["Выручка"] - pivot["Себестоимость"]
            pivot = pivot.sort_values(pivot.columns[0], ascending=False)
            st.dataframe(pivot.style.format("{:,.0f}"), use_container_width=True)

    clean = fdf.drop(columns=["_val", "_period"], errors="ignore")
    dc = st.columns(3)
    dc[0].download_button("Отфильтрованные данные (CSV)",
                          clean.to_csv(index=False).encode("utf-8-sig"),
                          "filtered_data.csv", "text/csv", use_container_width=True)
    with st.spinner("Формирование файла…"):
        xb = _xlsx_bytes(clean)
    dc[1].download_button("Отфильтрованные данные (XLSX)", xb,
                          "filtered_data.xlsx",
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                          use_container_width=True)
    if pivot is not None:
        dc[2].download_button("Сводная таблица (CSV)",
                              pivot.to_csv().encode("utf-8-sig"),
                              "summary.csv", "text/csv", use_container_width=True)

    # ════════ РАСШИРЕННЫЙ АНАЛИЗ ════════
    st.divider()
    st.markdown("#### Расширенный анализ")
    sub = st.tabs(["ABC / XYZ", "Сравнение периодов", "Аномалии", "Прогноз", "Отчёты"])
    with sub[0]:
        _render_abc_xyz(fdf, name_col, short_label, unit)
    with sub[1]:
        _render_compare(fdf, cat_col, short_label)
    with sub[2]:
        _render_anomalies(fdf, name_col, short_label)
    with sub[3]:
        _render_forecast(fdf, short_label, unit)
    with sub[4]:
        st.markdown("##### Готовые отчёты")
        st.caption("Выгрузка дашборда в единый файл для руководства.")
        rc = st.columns(2)
        with st.spinner("Формирование Excel-отчёта…"):
            xls = _excel_report_bytes(fdf, cat_col, name_col, short_label)
        rc[0].download_button("Excel-отчёт (несколько листов)", xls,
                              "report.xlsx",
                              "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                              type="primary", use_container_width=True, key="dlf_xlsreport")
        if rc[1].button("Сформировать PDF-отчёт", use_container_width=True, key="mv_pdfbtn"):
            try:
                with st.spinner("Рендеринг PDF…"):
                    st.session_state["pdf_report"] = _pdf_report_bytes(
                        fdf, cat_col, name_col, short_label, unit)
            except Exception as e:
                st.error(f"Не удалось сформировать PDF: {e}")
        if st.session_state.get("pdf_report"):
            rc[1].download_button("Скачать PDF-отчёт", st.session_state["pdf_report"],
                                  "report.pdf", "application/pdf",
                                  use_container_width=True, key="dlf_pdfreport")


def read_table_buffer(up):
    """Читает загруженный через uploader файл."""
    suf = Path(up.name).suffix.lower()
    if suf in (".xlsx", ".xls"):
        return pd.read_excel(up)
    sep = "\t" if suf == ".tsv" else None
    raw = up.getvalue().decode("utf-8", errors="replace")
    if sep is None:
        sep = ";" if raw[:4096].count(";") > raw[:4096].count(",") else ","
    return pd.read_csv(io.StringIO(raw), sep=sep)


# ══════════════════════════════════════════════════════════════════════════════
# TAB: ФАЙЛЫ
# ══════════════════════════════════════════════════════════════════════════════

def _list_files(folder):
    p = Path(folder)
    if not p.exists():
        return []
    return sorted(f for f in p.iterdir()
                  if f.is_file() and not f.name.startswith("."))


def tab_files(cfg):
    st.markdown("#### Управление файлами")

    # ── выбор каталогов прямо здесь (синхронно с настройками) ──
    with st.expander("Каталоги", expanded=False):
        st.caption("Изменение каталога сохраняется в настройках автоматически.")
        fc = st.columns(3)
        with fc[0]:
            cfg["input_folder"] = _folder_select("Источник", cfg["input_folder"], "ft_in")
        with fc[1]:
            cfg["output_folder"] = _folder_select("Результаты", cfg["output_folder"], "ft_out")
        with fc[2]:
            cfg["archive_folder"] = _folder_select("Архив", cfg["archive_folder"], "ft_arch")
        save_config(cfg)

    folders = {
        "Источник":   cfg["input_folder"],
        "Результаты": cfg["output_folder"],
        "Архив":      cfg["archive_folder"],
    }
    for f in folders.values():
        Path(f).mkdir(parents=True, exist_ok=True)

    # ── загрузка ──
    in_dir = Path(cfg["input_folder"])
    st.markdown(f"**Загрузка в источник** · `{cfg['input_folder']}/`")
    ups = st.file_uploader("XLSX, CSV или TSV", type=["xlsx", "xls", "csv", "tsv"],
                           accept_multiple_files=True, label_visibility="collapsed")
    if ups:
        for up in ups:
            (in_dir / up.name).write_bytes(up.getbuffer())
        st.success(f"Загружено файлов: {len(ups)}"); st.rerun()

    def _selkey(label, fname):
        return f"sel::{label}::{fname}"

    def _move_files(items, target_label):
        """items: список (label, folder, fname). Переносит в папку target_label."""
        target_folder = folders[target_label]
        moved = 0
        for label, folder, fname in items:
            if label == target_label:
                continue
            src = Path(folder) / fname
            dst = Path(target_folder) / fname
            try:
                if not src.exists():
                    continue
                if dst.exists():
                    stem, suf, i = dst.stem, dst.suffix, 1
                    while dst.exists():
                        dst = dst.with_name(f"{stem}_{i}{suf}"); i += 1
                shutil.move(str(src), str(dst))
                st.session_state.pop(_selkey(label, fname), None)
                moved += 1
            except Exception as e:
                st.error(f"«{fname}»: {e}")
        return moved

    # ── 1. ВЫБОР ФАЙЛОВ (галочки) ──
    st.divider()
    st.markdown("##### Выбор файлов")
    st.caption("Отметьте файлы галочками. Их можно перенести перетаскиванием ниже, "
               "скачать архивом или удалить.")
    cols = st.columns(len(folders))
    for col, (label, folder) in zip(cols, folders.items()):
        files = _list_files(folder)
        with col:
            st.markdown(
                f"<div class='folder-card'><span class='fc-title'>{label}</span>"
                f"<div class='fc-sub'>{folder}/ · файлов: {len(files)}</div></div>",
                unsafe_allow_html=True,
            )
            sc = st.columns(2)
            if sc[0].button("Выбрать все", key=f"selall_{label}",
                            use_container_width=True, disabled=not files):
                for f in files:
                    st.session_state[_selkey(label, f.name)] = True
                st.rerun()
            if sc[1].button("Снять выбор", key=f"desel_{label}",
                            use_container_width=True, disabled=not files):
                for f in files:
                    st.session_state[_selkey(label, f.name)] = False
                st.rerun()
            if not files:
                st.caption("Каталог пуст.")
                continue
            for f in files:
                kb = max(f.stat().st_size // 1024, 1)
                st.checkbox(f"{f.name}  ·  {kb} КБ", key=_selkey(label, f.name))

    # собрать выбранное
    selected = []   # (label, folder, fname)
    for label, folder in folders.items():
        for f in _list_files(folder):
            if st.session_state.get(_selkey(label, f.name)):
                selected.append((label, folder, f.name))

    # ── 2. ПЕРЕТАСКИВАНИЕ МЫШЬЮ ──
    st.divider()
    st.markdown("##### Перетаскивание в папку")
    st.caption("Перетащите файл мышью в другую колонку. Если отмечено несколько файлов — "
               "перетащив любой из них, вы перенесёте сразу все выбранные.")

    SEP = " ▸ "
    if sort_items is None:
        st.info("Компонент перетаскивания недоступен. Используйте кнопку «Переместить» ниже.")
    else:
        containers = [{"header": label,
                       "items": [f"{label}{SEP}{f.name}" for f in _list_files(folder)]}
                      for label, folder in folders.items()]
        disk_sig = abs(hash(tuple(
            (lbl, f.name) for lbl, fo in folders.items() for f in _list_files(fo)
        )))
        arranged = sort_items(
            containers, multi_containers=True, direction="vertical",
            key=f"dndboard_{disk_sig}",
            custom_style="""
            .sortable-component{display:flex;gap:10px;}
            .sortable-container{flex:1;background:rgba(128,128,128,.05);
              border:1px dashed rgba(128,128,128,.3);border-radius:14px;padding:8px;min-height:140px;}
            .sortable-container-header{font-weight:700;padding:6px 8px;opacity:.85;text-align:center;}
            .sortable-item{background:rgba(99,102,241,.14);border:1px solid rgba(99,102,241,.3);
              border-radius:9px;padding:7px 11px;margin:6px 0;cursor:grab;font-size:.82rem;}
            .sortable-item:hover{background:rgba(99,102,241,.26);}
            """,
        )
        prev = {it: c["header"] for c in containers for it in c["items"]}
        now  = {it: c["header"] for c in arranged for it in c["items"]}
        moved_items = [(it, now[it]) for it in now if prev.get(it) and now[it] != prev[it]]
        if moved_items:
            target_label = moved_items[0][1]
            to_move = list(selected)
            for it, _ in moved_items:
                lbl, fname = it.split(SEP, 1)
                to_move.append((lbl, folders[lbl], fname))
            # уникализируем
            uniq = {(l, f, n) for (l, f, n) in to_move}
            n = _move_files(sorted(uniq), target_label)
            if n:
                st.toast(f"Перемещено в «{target_label}»: {n}")
            st.rerun()

    # ── 3. ДЕЙСТВИЯ С ВЫБРАННЫМИ ──
    st.divider()
    st.markdown("##### Действия с выбранными")
    st.caption(f"Выбрано файлов: **{len(selected)}**")
    ac = st.columns([2, 1, 1, 1])
    target = ac[0].selectbox("Переместить в каталог", list(folders), key="bulk_target")
    if ac[1].button("Переместить", key="mv_bulk", use_container_width=True,
                    disabled=not selected):
        n = _move_files(selected, target)
        st.toast(f"Перемещено файлов: {n}"); st.rerun()

    if selected:
        import zipfile
        zbuf = io.BytesIO()
        with zipfile.ZipFile(zbuf, "w", zipfile.ZIP_DEFLATED) as zf:
            for label, folder, fname in selected:
                p = Path(folder) / fname
                if p.exists():
                    zf.write(p, arcname=f"{label}/{fname}")
        ac[2].download_button("Скачать ZIP", zbuf.getvalue(), "files.zip",
                              "application/zip", key="dlf_bulkzip",
                              use_container_width=True)
    else:
        ac[2].button("Скачать ZIP", disabled=True, use_container_width=True)

    if ac[3].button("Удалить", key="del_bulk", use_container_width=True,
                    disabled=not selected):
        removed = 0
        for label, folder, fname in selected:
            p = Path(folder) / fname
            try:
                if p.exists():
                    p.unlink(); removed += 1
                st.session_state.pop(_selkey(label, fname), None)
            except Exception as e:
                st.error(f"«{fname}»: {e}")
        st.toast(f"Удалено файлов: {removed}"); st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# TAB: СОСТОЯНИЕ
# ══════════════════════════════════════════════════════════════════════════════

_CFG_DEFAULTS = {
    "input_folder": "data", "category_folder": "categories",
    "output_folder": "merged", "archive_folder": "archive",
    "categorize": True, "anonymize": False, "name_match": False,
    "aggregate_after_merge": False, "model": DEFAULT_MODEL,
}


def tab_state():
    st.markdown("#### Конфигурация")
    st.caption("Наглядное редактирование параметров. Каталоги создаются автоматически "
               "при сохранении.")

    cfg = load_config()
    for k, v in _CFG_DEFAULTS.items():
        cfg.setdefault(k, v)

    # ── визуальный редактор ──
    with st.form("cfg_form"):
        st.markdown("**Каталоги**")
        gc = st.columns(2)
        in_f  = gc[0].text_input("Таблицы для консолидации", cfg["input_folder"])
        cat_f = gc[1].text_input("Источник категорий", cfg["category_folder"])
        gc2 = st.columns(2)
        out_f  = gc2[0].text_input("Результаты", cfg["output_folder"])
        arch_f = gc2[1].text_input("Архив", cfg["archive_folder"])

        st.markdown("**Этапы обработки**")
        tc = st.columns(2)
        v_cat  = tc[0].toggle(
            "Категоризация", value=cfg["categorize"],
            help="Автоматически присваивает каждому товару целевую категорию "
                 "по справочнику и данным источника. Работает мгновенно, без ИИ.")
        v_anon = tc[1].toggle(
            "Обезличивание", value=cfg["anonymize"],
            help="Заменяет реальные названия магазинов, адреса, бренды, компании "
                 "и товары на сгенерированные, а цены и количества — на изменённые.")
        v_agg  = st.toggle(
            "Агрегация после консолидации", value=cfg["aggregate_after_merge"],
            help="Схлопывает одинаковые строки в одну, суммируя числовые метрики "
                 "(выручку, количество, себестоимость).")

        fc = st.columns([1, 1, 3])
        save = fc[0].form_submit_button("Сохранить", type="primary",
                                        use_container_width=True)
        reset = fc[1].form_submit_button("Сбросить", use_container_width=True)

    if save:
        new = {
            "input_folder": in_f.strip() or "data",
            "category_folder": cat_f.strip() or "categories",
            "output_folder": out_f.strip() or "merged",
            "archive_folder": arch_f.strip() or "archive",
            "categorize": v_cat, "anonymize": v_anon,
            "aggregate_after_merge": v_agg,
        }
        for f in (new["input_folder"], new["category_folder"],
                  new["output_folder"], new["archive_folder"]):
            try:
                Path(f).mkdir(parents=True, exist_ok=True)
            except Exception:
                pass
        # применяется в начале следующего прогона (синхронизирует все виджеты)
        st.session_state["_apply_cfg"] = new
        st.toast("Конфигурация сохранена"); st.rerun()
    if reset:
        st.session_state["_apply_cfg"] = dict(_CFG_DEFAULTS)
        st.toast("Сброшено к значениям по умолчанию"); st.rerun()

    # текущее состояние
    with st.expander("Текущая конфигурация (config.json)"):
        st.json(cfg)

    st.divider()

    # ── редактор целевых категорий ──
    st.markdown("#### Целевые категории")
    cats = cm.load_target_categories()
    cat_text = st.text_area(f"По одной в строке (сейчас: {len(cats)})",
                            "\n".join(cats), height=200,
                            key=f"state_cats_{abs(hash(tuple(cats)))}")
    ec = st.columns([1, 1, 3])
    if ec[0].button("Сохранить категории", type="primary", use_container_width=True,
                    key="mv_save_cats"):
        new = [c.strip() for c in cat_text.splitlines() if c.strip()]
        _save_json(TARGET_CATS_PATH, new)
        st.toast(f"Сохранено категорий: {len(new)}"); st.rerun()
    if ec[1].button("Импортировать из источника", use_container_width=True,
                    key="mv_import_cats"):
        found = _suggest_categories(cfg["category_folder"])
        if found:
            _save_json(TARGET_CATS_PATH, found)
            st.toast(f"Импортировано: {len(found)}"); st.rerun()
        else:
            st.warning("В источнике категорий не найдено подходящих столбцов.")

    st.divider()

    # ── расширенное: raw JSON ──
    st.markdown("#### Служебные файлы (JSON)")
    state_files = {
        "category_overrides.json":  OVERRIDES_PATH,
        "category_cache.json":      CACHE_PATH,
        "name_matches.json":        NAME_MATCHES_PATH,
        "analytics_presets.json":   PRESETS_PATH,
    }
    for label, path in state_files.items():
        data = _read_json(path, None)
        size = len(data) if isinstance(data, (list, dict)) else 0
        exists = Path(path).exists()
        with st.expander(f"`{label}` " + (f"· записей: {size}" if exists else "· файл отсутствует")):
            raw = json.dumps(data, ensure_ascii=False, indent=2) if exists else "{}"
            new_raw = st.text_area(label, raw, height=220,
                                   key=f"st_{label}_{abs(hash(raw))}",
                                   label_visibility="collapsed")
            if st.button("Сохранить", key=f"mv_save_{label}", type="primary"):
                try:
                    _save_json(path, json.loads(new_raw))
                    st.toast(f"Файл сохранён: {label}"); st.rerun()
                except json.JSONDecodeError as e:
                    st.error(f"Ошибка формата JSON: {e}")

    st.divider()
    st.markdown("#### Сброс кэшей")
    cc = st.columns(3)
    if cc[0].button("Очистить кэш категоризации", key="warn_clr_cache",
                    use_container_width=True):
        _save_json(CACHE_PATH, {}); st.toast("Кэш категоризации очищен")
    if cc[1].button("Очистить ручные правки", key="warn_clr_ov",
                    use_container_width=True):
        _save_json(OVERRIDES_PATH, {}); st.toast("Ручные правки очищены")
    if cc[2].button("Очистить правила сопоставления", key="warn_clr_nm",
                    use_container_width=True):
        _save_json(NAME_MATCHES_PATH, {}); st.toast("Правила сопоставления очищены")


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR (глобальные настройки → config.json)
# ══════════════════════════════════════════════════════════════════════════════

def sidebar_config():
    cfg = load_config()
    # дефолты на случай отсутствия ключей
    cfg.setdefault("input_folder", "data")
    cfg.setdefault("category_folder", "categories")
    cfg.setdefault("output_folder", "merged")
    cfg.setdefault("archive_folder", "archive")
    cfg.setdefault("anonymize", False)
    cfg.setdefault("categorize", True)
    cfg.setdefault("name_match", False)
    cfg.setdefault("aggregate_after_merge", False)
    cfg.setdefault("model", DEFAULT_MODEL)

    with st.sidebar:
        st.markdown("### Настройки")

        st.markdown("**Оформление**")
        pal = st.selectbox("Цветовая тема", list(PALETTES),
                           index=list(PALETTES).index(st.session_state.get("ui_palette", "Индиго")),
                           key="ui_palette_select")
        if pal != st.session_state.get("ui_palette"):
            st.session_state["ui_palette"] = pal
            st.rerun()

        st.divider()
        st.markdown("**Каталоги**")
        st.caption("Выберите существующий каталог или создайте новый — он будет создан автоматически.")
        cfg["input_folder"]    = _folder_select("Таблицы для консолидации", cfg["input_folder"], "sb_in")
        cfg["category_folder"] = _folder_select("Источник категорий", cfg["category_folder"], "sb_cat")
        cfg["output_folder"]   = _folder_select("Результаты", cfg["output_folder"], "sb_out")
        cfg["archive_folder"]  = _folder_select("Архив", cfg["archive_folder"], "sb_arch")

        st.divider()
        st.markdown("**Этапы обработки**")
        st.caption("Категоризация и обезличивание переключаются на вкладке «Обработка».")
        cfg["aggregate_after_merge"] = st.toggle(
            "Агрегация после консолидации", value=cfg["aggregate_after_merge"],
            key="sb_aggregate",
            help="Объединение строк с суммированием метрик")

        st.divider()
        if st.button("Сохранить настройки", type="primary", use_container_width=True):
            save_config(cfg)
            st.toast("Настройки сохранены")
        # автосохранение текущих значений на каждый рендер, чтобы пайплайн видел флаги
        save_config(cfg)

    return cfg


# ══════════════════════════════════════════════════════════════════════════════
# ENTRY
# ══════════════════════════════════════════════════════════════════════════════

st.set_page_config(page_title="EnigmaSQL", layout="wide", page_icon="🧠")

# ── ТЕМЫ ОФОРМЛЕНИЯ ──
PALETTES = {
    "Индиго":  ("#6366F1", "#8B5CF6", "#22D3EE"),
    "Изумруд": ("#10B981", "#059669", "#34D399"),
    "Океан":   ("#0EA5E9", "#2563EB", "#22D3EE"),
    "Графит":  ("#475569", "#334155", "#64748B"),
    "Янтарь":  ("#F59E0B", "#D97706", "#FBBF24"),
}


def _hex_rgb(h):
    h = h.lstrip("#")
    return tuple(int(h[i:i + 2], 16) for i in (0, 2, 4))


def inject_css(palette_name="Индиго"):
    a1, a2, a3 = PALETTES.get(palette_name, PALETTES["Индиго"])
    r, g, b = _hex_rgb(a1)
    glow = f"{r},{g},{b}"
    st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

:root{{
  --accent:{a1}; --accent2:{a2}; --accent3:{a3}; --glow:{glow};
  --card-radius:16px;
}}

html, body, .stApp, [class*="css"]{{font-family:'Inter',system-ui,sans-serif;}}
#MainMenu,[data-testid="stToolbar"],footer,[data-testid="stDecoration"]{{display:none;}}

/* ── АДАПТИВНЫЙ КОНТЕЙНЕР ── */
.block-container{{
  padding-top:1.1rem; padding-bottom:3rem;
  max-width:min(1360px, 96vw);
  padding-left:clamp(.6rem, 2vw, 3rem);
  padding-right:clamp(.6rem, 2vw, 3rem);
}}

/* ── ГРАДИЕНТНАЯ ШАПКА ── */
.enigma-hero{{
  background:linear-gradient(110deg,var(--accent) 0%,var(--accent2) 55%,var(--accent3) 130%);
  border-radius:var(--card-radius);
  padding:clamp(16px,2.5vw,26px) clamp(18px,3vw,30px); margin-bottom:18px;
  box-shadow:0 10px 30px -10px rgba(var(--glow),.55);
}}
.enigma-hero h1{{
  color:#fff; font-size:clamp(1.3rem,3vw,1.8rem); font-weight:800;
  margin:0; letter-spacing:-.5px;
}}
.enigma-hero p{{color:rgba(255,255,255,.88); margin:.3rem 0 0; font-size:clamp(.8rem,1.4vw,.95rem);}}

/* ── ВКЛАДКИ (пилюли) ── */
[data-testid="stTabs"] [data-baseweb="tab-list"]{{
  gap:6px; background:rgba(128,128,128,.08); padding:6px;
  border-radius:14px; border:1px solid rgba(128,128,128,.12);
  flex-wrap:wrap;
}}
[data-testid="stTabs"] [data-baseweb="tab-list"] button[role="tab"]{{
  height:40px; border-radius:10px; padding:0 clamp(10px,1.5vw,18px);
  font-weight:600; background:transparent; border:none; transition:all .18s ease;
}}
[data-testid="stTabs"] [data-baseweb="tab-list"] button[role="tab"]:hover{{
  background:rgba(var(--glow),.14);
}}
[data-testid="stTabs"] [data-baseweb="tab-list"] button[aria-selected="true"]{{
  background:linear-gradient(120deg,var(--accent),var(--accent2));
  color:#fff !important; box-shadow:0 4px 14px -4px rgba(var(--glow),.6);
}}
[data-testid="stTabs"] [data-baseweb="tab-highlight"],
[data-testid="stTabs"] [data-baseweb="tab-border"]{{display:none;}}

/* ── МЕТРИКИ-КАРТОЧКИ ── */
[data-testid="stMetric"]{{
  background:rgba(128,128,128,.07); border:1px solid rgba(128,128,128,.14);
  border-radius:14px; padding:14px 16px;
  transition:transform .15s ease, box-shadow .15s ease;
}}
[data-testid="stMetric"]:hover{{
  transform:translateY(-2px); box-shadow:0 8px 20px -8px rgba(var(--glow),.4);
  border-color:rgba(var(--glow),.4);
}}
[data-testid="stMetricValue"]{{font-weight:800; font-size:clamp(1.1rem,2vw,1.5rem);}}
[data-testid="stMetricLabel"]{{opacity:.75; font-weight:600;}}

/* ── КНОПКИ ── */
.stButton>button, .stDownloadButton>button{{
  border-radius:10px; font-weight:600; border:1px solid rgba(128,128,128,.2);
  transition:all .15s ease;
}}
.stButton>button:hover, .stDownloadButton>button:hover{{
  transform:translateY(-1px); border-color:var(--accent);
}}
.stButton>button[kind="primary"], .stDownloadButton>button[kind="primary"]{{
  background:linear-gradient(120deg,var(--accent),var(--accent2));
  border:none; color:#fff; box-shadow:0 4px 14px -4px rgba(var(--glow),.6);
}}

/* цветные кнопки по префиксу ключа (Streamlit .st-key-*) */
[class*="st-key-del_"] button{{background:linear-gradient(120deg,#DC2626,#EF4444)!important;border:none!important;color:#fff!important;}}
[class*="st-key-dlf_"] button{{background:linear-gradient(120deg,#16A34A,#22C55E)!important;border:none!important;color:#fff!important;}}
[class*="st-key-mv_"]  button{{background:linear-gradient(120deg,var(--accent),var(--accent2))!important;border:none!important;color:#fff!important;}}
[class*="st-key-warn_"] button{{background:linear-gradient(120deg,#D97706,#F59E0B)!important;border:none!important;color:#fff!important;}}

/* ── ЭКСПАНДЕРЫ ── */
[data-testid="stExpander"]{{
  border:1px solid rgba(128,128,128,.16); border-radius:14px;
  overflow:hidden; background:rgba(128,128,128,.04);
}}
[data-testid="stExpander"] summary{{font-weight:600;}}

[data-baseweb="select"]>div, .stTextInput input, .stNumberInput input{{
  border-radius:10px !important;
}}

/* ── САЙДБАР ── */
[data-testid="stSidebar"]{{
  background:linear-gradient(180deg,rgba(var(--glow),.10),rgba(var(--glow),.02));
  border-right:1px solid rgba(128,128,128,.12);
}}
[data-testid="stSidebar"] .stButton>button{{width:100%;}}

[data-testid="stDataFrame"], [data-testid="stTable"]{{
  border-radius:12px; overflow:hidden; border:1px solid rgba(128,128,128,.14);
}}
[data-testid="stVegaLiteChart"]{{
  background:rgba(128,128,128,.04); border-radius:14px; padding:10px;
  border:1px solid rgba(128,128,128,.1);
}}

hr{{margin:1.1rem 0; border-color:rgba(128,128,128,.15);}}
h5{{font-weight:700; letter-spacing:-.2px;}}

/* ── карточка файла / папки ── */
.folder-card{{
  border:1px solid rgba(128,128,128,.16); border-radius:14px;
  padding:12px 14px; margin-bottom:8px;
  background:linear-gradient(135deg,rgba(var(--glow),.06),rgba(128,128,128,.03));
}}
.folder-card .fc-title{{font-weight:700; font-size:.95rem;}}
.folder-card .fc-sub{{opacity:.65; font-size:.78rem;}}
.file-chip{{
  display:inline-block; padding:3px 10px; margin:3px 4px 3px 0;
  border-radius:8px; background:rgba(var(--glow),.12);
  border:1px solid rgba(var(--glow),.25); font-size:.82rem;
}}

/* ── МАСШТАБИРОВАНИЕ ПОД МАЛЫЕ ЭКРАНЫ ── */
@media (max-width: 900px){{
  [data-testid="stSidebar"]{{min-width:240px;}}
  .enigma-hero{{padding:14px 16px;}}
}}
@media (max-width: 640px){{
  [data-testid="stTabs"] [data-baseweb="tab-list"] button[role="tab"]{{
    padding:0 9px; font-size:.82rem; height:36px;
  }}
}}
</style>
""", unsafe_allow_html=True)


inject_css(st.session_state.get("ui_palette", "Индиго"))

# применяем конфигурацию, сохранённую на вкладке «Состояние», ДО создания виджетов,
# чтобы синхронизировать сайдбар, переключатели и поля выбора каталогов
if not _pipe_err and "_apply_cfg" in st.session_state:
    _pend = st.session_state.pop("_apply_cfg")
    _base = load_config(); _base.update(_pend); save_config(_base)
    for _f in (_pend.get("input_folder"), _pend.get("category_folder"),
               _pend.get("output_folder"), _pend.get("archive_folder")):
        if _f:
            try:
                Path(_f).mkdir(parents=True, exist_ok=True)
            except Exception:
                pass
    _sync = {
        "sb_in": _pend.get("input_folder"),   "ft_in": _pend.get("input_folder"),
        "sb_out": _pend.get("output_folder"), "ft_out": _pend.get("output_folder"),
        "sb_arch": _pend.get("archive_folder"), "ft_arch": _pend.get("archive_folder"),
        "sb_cat": _pend.get("category_folder"),
        "run_categorize": _pend.get("categorize"),
        "run_anonymize": _pend.get("anonymize"),
        "sb_aggregate": _pend.get("aggregate_after_merge"),
    }
    for _k, _v in _sync.items():
        if _v is not None:
            st.session_state[_k] = _v

st.markdown("""
<div class="enigma-hero">
  <h1>EnigmaSQL</h1>
  <p>Платформа обработки данных: консолидация таблиц, категоризация, обезличивание и аналитика продаж</p>
</div>
""", unsafe_allow_html=True)

if _pipe_err:
    st.error(f"Не удалось загрузить модули обработки:\n\n```\n{_pipe_err}\n```\n\n"
             "Убедитесь, что `pipeline.py` и `category_matcher.py` находятся рядом с приложением.")
else:
    cfg = sidebar_config()
    t_run, t_an, t_cat, t_files, t_state = st.tabs(
        ["Обработка", "Аналитика", "Категории", "Файлы", "Состояние"]
    )
    with t_run:
        tab_run(cfg)
    with t_an:
        tab_analytics(cfg)
    with t_cat:
        tab_categorize(cfg)
    with t_files:
        tab_files(cfg)
    with t_state:
        tab_state()
