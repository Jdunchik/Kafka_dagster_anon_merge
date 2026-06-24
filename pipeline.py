import re
import sys
import json
import time
import hashlib
import random
import difflib
import shutil
from collections.abc import Callable
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# ── CONSTANTS ──────────────────────────────────────────────────────────────────

CONFIG_PATH   = Path("config.json")
MAPPINGS_PATH = Path("mappings.json")
PRICE_COEF     = 1.117
QTY_SHIFT      = 2
SUPPORTED_EXTS = {".xlsx", ".xls", ".xlsm", ".xlsb", ".csv", ".tsv", ".parquet"}

CANONICAL = {
    "week":                "keep",
    "month":               "keep",
    "year":                "keep",
    "store_code":          "drop",
    "store_name":          "anon_store",
    "address":             "anon_store",
    "distribution_center": "anon_company",
    "region":              "keep",
    "store_format":        "keep",
    "store_subformat":     "keep",
    "category_0":          "keep",
    "category_1":          "keep",
    "category_2":          "keep",
    "category_3":          "keep",
    "category_4":          "keep",
    "item_code":           "drop",
    "item_name":           "anon_product",
    "brand":               "anon_brand",
    "manufacturer":        "anon_company",
    "supplier":            "anon_company",
    "weight":              "keep",
    "unit":                "keep",
    "barcode":             "drop",
    "qty_sold":            "add_qty",
    "sales_rub":           "mul_price",
    "cost_rub":            "mul_price",
}

DESIRED_ORDER = [
    "week", "month", "year",
    "store_name", "address", "region", "store_format", "store_subformat",
    "distribution_center",
    "category_0", "category_1", "category_2", "category_3", "category_4",
    "target_category", "category_confidence",
    "canonical_name", "item_name", "brand", "manufacturer", "supplier", "weight", "unit",
    "qty_sold", "sales_rub", "cost_rub",
]

ALIASES = {
    "week":                ["неделя", "нед", "week", "wk", "неделя продаж"],
    "month":               ["месяц", "мес", "month", "mo"],
    "year":                ["год", "year", "yr", "год продаж"],
    "store_code":          ["код", "код магазина", "код тт", "store_code",
                            "id магазина", "код торговой точки"],
    "store_name":          ["название магазина", "магазин", "маг","наим магазина",
                            "тт", "store", "store_name", "торговая точка"],
    "address":             ["адрес", "address", "addr", "адрес магазина"],
    "distribution_center": ["основной рц", "рц", "distribution center",
                            "distribution_center", "распределительный центр",
                            "основной распределительный центр", "склад", "dc"],
    "region":              ["регион", "region", "reg", "рег", "регион продаж",
                            "территория", "регион тт", "область"],
    "store_format":        ["формат", "format", "тип магазина",
                            "формат магазина", "формат тт"],
    "store_subformat":     ["субформат", "subformat", "sub-format",
                            "подформат", "суб-формат", "тип тт"],
    "category_0":          ["уровень 0", "уровень0", "level 0", "category_0",
                            "cat0", "lvl0", "категория 0"],
    "category_1":          ["уровень 1", "уровень1", "категория 1", "level 1",
                            "category_1", "cat1", "lvl1"],
    "category_2":          ["уровень 2", "уровень2", "категория 2", "level 2",
                            "category_2", "cat2", "lvl2"],
    "category_3":          ["уровень 3", "уровень3", "категория 3", "level 3",
                            "category_3", "cat3", "lvl3"],
    "category_4":          ["уровень 4", "уровень4", "категория 4", "level 4",
                            "category_4", "cat4", "lvl4",
                            "subject", "предмет", "категория", "category",
                            "group", "группа", "подкатегория"],
    "item_code":           ["код позиции", "код товара", "sku", "item_code",
                            "артикул", "арт", "код поз", "article",
                            "product_id", "product_code", "код продукта"],
    "item_name":           ["наименование", "название товара", "товар",
                            "позиция", "item_name", "item", "наим", "назв",
                            "name", "full_name", "short_name", "номенклатура",
                            "наименование номенклатуры", "товарная позиция",
                            "product_name", "goods_name", "sku_name",
                            "наименование товара", "название продукта",
                            "наименование продукта", "наим товара",
                            "наим продукта", "наименование позиции"],
    "brand":               ["бренд", "brand", "марка", "торговая марка", "brandname"],
    "manufacturer":        ["производитель", "произв", "пр-ль", "manufacturer", "mfr",
                            "producer", "изготовитель"],
    "supplier":            ["поставщик", "supplier", "spl", "поставщ",
                            "дистрибьютор"],
    "weight":              ["вес", "weight", "масса", "вес нетто",
                            "net weight", "вес брутто", "масса нетто"],
    "unit":                ["единица измерения", "ед. изм.", "ед.изм.", "unit",
                            "units", "ед изм", "единица", "ед"],
    "barcode":             ["штриховой код", "штрихкод", "ean", "barcode",
                            "bar code", "код штрих", "баркод"],
    "qty_sold":            ["продажи в шт", "продажи в шт.", "qty", "qty_sold",
                            "шт", "количество", "кол-во", "продажи в штуках",
                            "quantity", "qty_pcs", "sold_qty", "продано",
                            "продажи шт", "кол-во проданных", "продажи в натур",
                            "реализация шт", "отгрузка шт"],
    "sales_rub":           ["продажи в руб", "продажи в руб.", "sales_rub",
                            "выручка", "оборот", "продажи в рублях", "sales",
                            "revenue", "выручка от продаж", "реализация",
                            "выручка от реализации", "оборот в руб",
                            "сумма продаж", "реализация руб"],
    "cost_rub":            ["себестоимость", "себестоимсть",
                            "себестоимость в руб", "себестоимсть в руб",
                            "cost_rub", "cost", "себес", "закупка",
                            "закупочная цена"],
}


# ── WORD POOLS (используются только при anonymize=True) ─────────────────────────

_BRAND_PREFIXES = [
    "АРК","БЛЕ","ВЕЛ","ГРИ","ДИВ","ЗЕН","КЛИ","ЛЮМ","МИР","НОР",
    "ОМЕ","ПРИ","РАД","СВЕ","ТИП","УЛЬ","ФОР","ХОР","ЮГА","ЯРО",
    "АЛЬ","БОР","ВОЛ","ГАЛ","ДОН","ЖАР","ЗОР","ИВА","КАМ","ЛАД",
    "МАГ","НАД","ОРА","ПАЛ","РЯД","СИЛ","ТАЛ","УРА","ФАЛ","ЦЕН",
    "ЧИС","ШАЛ","ЩЕД","ЭЛЬ","ЮНА","ЯСЕ","АВА","БАЛ","ВАЛ","ГОР",
]
_BRAND_SUFFIXES = [
    "А","ЕКС","ИТА","ОН","ОС","УМ","ЕТ","АН","ЕЛ","ЕВА",
    "ИН","АР","ОР","ЕЖ","ОВА","ИМ","ЕР","АЛ","ИС","ЮМ",
]
_STORE_WORDS = [
    "Рябина","Берёза","Клён","Сосна","Ясень","Дуб","Каштан","Ива","Кедр","Тополь",
    "Заря","Рассвет","Весна","Маяк","Радуга","Горизонт","Меридиан","Сфера","Орбита",
    "Вектор","Импульс","Прогресс","Полюс","Сигма","Феникс","Атлант","Аврора","Ладья",
    "Родник","Исток","Утёс","Бриз","Прибой","Янтарь","Топаз","Рубин","Агат","Оникс",
    "Кварц","Малахит","Опал","Жемчуг","Сапфир","Изумруд","Гранит","Базальт","Мрамор",
    "Ветер","Буря","Гроза","Туман","Иней","Метель","Поток","Якорь","Штурвал","Компас",
    "Фрегат","Бриг","Колос","Росток","Лепесток","Бутон","Нептун","Борей","Зефир",
    "Дубрава","Роща","Поляна","Опушка","Апекс","Зенит","Вершина","Долина","Луга",
    "Курган","Равнина","Перекрёсток","Рубеж","Форт","Цитадель","Твердыня",
    "Авангард","Арьергард","Резерв","Запас","Клад","Сокровище","Копилка",
]

_DESCRIPTORS = {
    "посуд":     ["Средство для мытья посуды","Гель для посуды","Жидкость для посуды","Бальзам для посуды"],
    "чистящ":    ["Крем чистящий","Средство чистящее","Порошок чистящий","Паста чистящая"],
    "стирк":     ["Гель для стирки","Жидкость для стирки","Порошок стиральный","Капсулы для стирки"],
    "унитаза":   ["Блок для унитаза","Таблетка для унитаза","Гель для унитаза","Диск для унитаза"],
    "сантехник": ["Средство для сантехники","Очиститель сантехники","Гель для ванной"],
    "уборк":     ["Средство для уборки","Спрей для уборки","Универсальный очиститель"],
    "освеж":     ["Освежитель воздуха","Спрей-освежитель","Ароматизатор воздуха"],
    "ополаск":   ["Ополаскиватель","Кондиционер для белья","Бальзам-ополаскиватель"],
    "дезинфект": ["Дезинфицирующее средство","Антисептик","Дезинфектант"],
    "пятновы":   ["Пятновыводитель","Отбеливатель","Средство от пятен"],
}
_VARIANTS = [
    "Лаванда","Лимон","Морозная свежесть","Зелёный чай","Яблоко","Хвоя",
    "Ромашка","Океан","Мята","Роза","Кедр","Цитрус","Лайм","Ваниль",
    "Жасмин","Ландыш","Морской бриз","Альпийский луг","Утренняя роса","Хлопок",
]
_SIZES = [
    "500мл","750мл","1000мл","1.3кг","2.4кг","600г","3х50г","450мл","800мл",
    "1.5кг","400мл","2кг","1.2л","350мл","1800г","900мл","5л","200мл","300мл",
]


# ── GENERATORS (anonymize=True) ────────────────────────────────────────────────

def _seed(key: str) -> int:
    return int(hashlib.md5(str(key).encode("utf-8")).hexdigest()[:8], 16)

def _rng(key: str) -> random.Random:
    return random.Random(_seed(key))

def _faker(key: str):
    from faker import Faker as _Faker   # ленивый импорт: faker нужен только при анонимизации
    f = _Faker("ru_RU")
    f.seed_instance(_seed(key))
    return f

def gen_store(key: str, used_names: set) -> dict:
    r = _rng(key)
    f = _faker(key)
    base = r.choice(_STORE_WORDS)
    name, n = base, 1
    while name in used_names:
        name = base + str(n); n += 1
    addr = f"{f.region()}, {f.city()}, {f.street_name()}, {f.building_number()}"
    return {"name": name, "address": addr}

def gen_brand(key: str, used: set) -> str:
    r = _rng(key)
    base = r.choice(_BRAND_PREFIXES) + r.choice(_BRAND_SUFFIXES)
    cand, n = base, 1
    while cand in used:
        cand = base + str(n); n += 1
    return cand

def gen_company(key: str, used: set = None) -> str:
    f = _faker(key)
    base = f.company()
    if used is None:
        return base
    cand, n = base, 1
    while cand in used:
        cand = f"{base} {n}"; n += 1
    return cand

def gen_product(key: str, category: str, fake_brand: str) -> str:
    r = _rng(key)
    cat = str(category).lower()
    descriptor = next(
        (r.choice(opts) for kw, opts in _DESCRIPTORS.items() if kw in cat),
        "Средство",
    )
    return f"{fake_brand} {descriptor} {r.choice(_VARIANTS)} {r.choice(_SIZES)}"


# ── FILE I/O ───────────────────────────────────────────────────────────────────

def read_table(path: Path) -> pd.DataFrame:
    ext = path.suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        return pd.read_excel(path)
    elif ext == ".xls":
        return pd.read_excel(path, engine="xlrd")
    elif ext == ".xlsb":
        return pd.read_excel(path, engine="pyxlsb")
    elif ext in (".csv", ".tsv"):
        sep = "\t" if ext == ".tsv" else None   # None → sniff below
        for enc in ("utf-8-sig", "cp1251", "latin-1"):
            try:
                raw = path.read_bytes()
                raw.decode(enc)                 # probe
                if sep is None:
                    sample = raw.decode(enc)[:4096]
                    sep = ";" if sample.count(";") > sample.count(",") else ","
                return pd.read_csv(path, sep=sep, encoding=enc)
            except (UnicodeDecodeError, ValueError):
                continue
        raise ValueError(f"Не удалось определить кодировку: {path.name}")
    elif ext == ".parquet":
        return pd.read_parquet(path)
    raise ValueError(f"Неподдерживаемый формат: {ext}")

def _format_xlsx(path: Path):
    wb = load_workbook(path)
    ws = wb.active
    if ws is None:
        return
    header_fill = PatternFill("solid", start_color="1F4E79")
    for cell in ws[1]:
        cell.font      = Font(bold=True, name="Arial", size=10, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 38
    ws.freeze_panes = "A2"
    for ci in range(1, ws.max_column + 1):
        col_letter = get_column_letter(ci)
        maxlen = max(
            (len(str(ws.cell(row=r, column=ci).value or ""))
             for r in range(1, min(ws.max_row + 1, 200))),
            default=8,
        )
        ws.column_dimensions[col_letter].width = min(maxlen + 3, 52)
    wb.save(path)

def write_xlsx(df: pd.DataFrame, path: Path):
    df.to_excel(path, index=False)
    _format_xlsx(path)


# ── MAPPINGS (anonymize=True) ──────────────────────────────────────────────────

def load_mappings() -> dict:
    defaults = {"stores": {}, "brands": {}, "companies": {}, "products": {}}
    if not MAPPINGS_PATH.exists():
        return defaults
    data = json.loads(MAPPINGS_PATH.read_text(encoding="utf-8"))
    if "companies" not in data:
        data["companies"] = {**data.pop("manufacturers", {}), **data.pop("suppliers", {})}
    return {**defaults, **data}

def save_mappings(m: dict):
    MAPPINGS_PATH.write_text(json.dumps(m, ensure_ascii=False, indent=2), encoding="utf-8")

def build_mappings(df: pd.DataFrame, col: dict, m: dict, log_fn: Callable[[str], None]) -> dict:
    key_col = col.get("store_code") or col.get("store_name")
    used_names = {v["name"] for v in m["stores"].values()}
    new_s = 0
    if key_col and key_col in df.columns:
        for val in sorted(df[key_col].dropna().unique()):
            k = str(int(val)) if isinstance(val, float) else str(val).strip()
            if k not in m["stores"]:
                store = gen_store(k, used_names)
                m["stores"][k] = store
                used_names.add(store["name"])
                new_s += 1
    log_fn(f"  Магазины: +{new_s} новых (кэш: {len(m['stores'])})")

    bc = col.get("brand")
    used_brands = set(m["brands"].values())
    new_b = 0
    if bc and bc in df.columns:
        for brand in sorted(df[bc].dropna().unique()):
            k = str(brand).strip()
            if k not in m["brands"]:
                m["brands"][k] = gen_brand(k, used_brands)
                used_brands.add(m["brands"][k])
                new_b += 1
    log_fn(f"  Бренды: +{new_b} новых (кэш: {len(m['brands'])})")

    new_c = 0
    used_co = set(m["companies"].values())
    for ckey in ["manufacturer", "supplier", "distribution_center"]:
        c = col.get(ckey)
        if c and c in df.columns:
            for val in sorted(df[c].dropna().unique()):
                k = str(val).strip()
                if k not in m["companies"]:
                    m["companies"][k] = gen_company(k, used_co)
                    used_co.add(m["companies"][k])
                    new_c += 1
    log_fn(f"  Компании: +{new_c} новых (кэш: {len(m['companies'])})")

    ic   = col.get("item_name")
    cat2 = col.get("category_2")
    bc2  = col.get("brand")
    new_p = 0
    if ic and ic in df.columns:
        needed = [c for c in [ic, bc2, cat2] if c and c in df.columns]
        subset = df[needed].drop_duplicates().dropna(subset=[ic])
        subset = subset[~subset[ic].astype(str).isin(m["products"])]
        for _, row in subset.iterrows():
            k  = str(row[ic]).strip()
            fb = m["brands"].get(str(row.get(bc2, "")).strip(), "БРЕНД") if bc2 else "БРЕНД"
            cat = str(row.get(cat2, "")) if cat2 else ""
            m["products"][k] = gen_product(k, cat, fb)
            new_p += 1
    log_fn(f"  Товары: +{new_p} новых (кэш: {len(m['products'])})")
    return m

def _normalize(col: str) -> str:
    s = col.lower().strip()
    for suffix in [" с ндс", " без ндс", ". с ндс", " с nds", " incl. vat", " excl. vat"]:
        if s.endswith(suffix):
            s = s[:-len(suffix)]
    s = s.rstrip(". ,")
    # убрать всю пунктуацию и лишние пробелы для сравнения
    s = re.sub(r"[^\w\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _alias_norm(s: str) -> str:
    """Нормализация алиаса: нижний регистр, '_' → пробел, без пунктуации."""
    s = s.lower().replace("_", " ")
    s = re.sub(r"[^\w\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _col_score(col_norm: str, aliases: list) -> float:
    """
    Score совпадения нормализованного имени колонки с одним из алиасов.

    Уровни (max):
      1. Точное совпадение        → 1.0
      2. Алиас ⊆ col (подстрока)  → 0.60–1.0 (пропорц. длине алиаса)
         Только алиасы длиной ≥ 5 символов, чтобы избежать "ed"⊆"name"
      3. Token-set Jaccard        → 0.0–1.0

    НЕ используем «col ⊆ алиас»: "name" ⊆ "store_name" давало ложные матчи.
    """
    col_toks = set(col_norm.split())
    best = 0.0
    for alias in aliases:
        a = _alias_norm(alias)
        if not a:
            continue
        if col_norm == a:
            return 1.0
        # алиас является подстрокой названия колонки
        # требуем покрытие > 50%: "quantity" (8) в "packaging quantity" (19) = 42% → нет
        if len(a) >= 5 and a in col_norm:
            coverage = len(a) / max(len(col_norm), 1)
            if coverage >= 0.50:
                best = max(best, 0.6 + 0.4 * coverage)
        # Jaccard по токенам
        a_toks = set(a.split())
        if a_toks and col_toks:
            j = len(col_toks & a_toks) / len(col_toks | a_toks)
            best = max(best, j)
    return best


def fuzzy_map(actual_cols: list) -> dict:
    """
    {actual_col → canonical_field}

    Проход 1 — точный алиас: каждая колонка сравнивается со всеми алиасами
      построчно.  Колонка, попавшая в точный матч, больше НЕ участвует
      в нечётком поиске (даже если «своё» поле уже занято).

    Проход 2 — fuzzy (только колонки без точного алиаса):
      _col_score ≥ 0.60, жадное назначение по убыванию score.
    """
    import heapq

    alias_to_canon = {_alias_norm(alias): cname
                      for cname, aliases in ALIASES.items()
                      for alias in aliases}

    result: dict   = {}
    used_canon     = set()
    exact_cols     = set()   # колонки с точным алиасом (даже если поле занято)

    # Проход 1: точный матч
    for col in actual_cols:
        norm = _normalize(col)
        if norm in alias_to_canon:
            exact_cols.add(col)
            cname = alias_to_canon[norm]
            if cname not in used_canon:
                result[col] = cname
                used_canon.add(cname)

    # Проход 2: fuzzy только для колонок без точного алиаса
    scores: dict = {}
    for col in actual_cols:
        if col in result or col in exact_cols:
            continue
        norm = _normalize(col)
        cands = []
        for cname, aliases in ALIASES.items():
            if cname in used_canon:
                continue
            s = _col_score(norm, aliases)
            if s >= 0.60:
                cands.append((s, cname))
        if cands:
            scores[col] = sorted(cands, reverse=True)

    heap = []
    for col, cands in scores.items():
        if cands:
            heapq.heappush(heap, (-cands[0][0], col, cands[0][1]))

    while heap:
        neg_s, col, cname = heapq.heappop(heap)
        if col in result or cname in used_canon:
            remaining = [(s, c) for s, c in scores.get(col, [])
                         if c not in used_canon and col not in result]
            if remaining:
                heapq.heappush(heap, (-remaining[0][0], col, remaining[0][1]))
            continue
        result[col] = cname
        used_canon.add(cname)

    return result


def get_col_map(df: pd.DataFrame) -> dict:
    raw = fuzzy_map(list(df.columns))
    canon_to_actual: dict = {}
    for act_col, cname in raw.items():
        if cname and cname not in canon_to_actual:
            canon_to_actual[cname] = act_col
    return canon_to_actual


# ── TRANSFORM ──────────────────────────────────────────────────────────────────

def transform(df: pd.DataFrame, col: dict, m: dict, anonymize: bool = True) -> pd.DataFrame:
    if anonymize:
        sc   = col.get("store_code")
        sn_c = col.get("store_name")
        sa_c = col.get("address")
        key_col = sc if (sc and sc in df.columns) else sn_c

        if key_col and key_col in df.columns:
            store_name_map: dict = {}
            store_addr_map: dict = {}
            for k, v in m["stores"].items():
                try:
                    store_name_map[int(k)] = v["name"]
                    store_addr_map[int(k)] = v["address"]
                except ValueError:
                    store_name_map[k] = v["name"]
                    store_addr_map[k] = v["address"]
            if sn_c and sn_c in df.columns:
                df[sn_c] = df[key_col].map(store_name_map)
            if sa_c and sa_c in df.columns:
                df[sa_c] = df[key_col].map(store_addr_map)

        for cname, action in CANONICAL.items():
            actual = col.get(cname)
            if not actual or actual not in df.columns:
                continue
            if action in ("anon_store", "drop"):
                continue
            elif action == "anon_brand":
                df[actual] = df[actual].astype(str).str.strip().map(m["brands"])
            elif action == "anon_company":
                df[actual] = df[actual].astype(str).str.strip().map(m["companies"])
            elif action == "anon_product":
                df[actual] = df[actual].astype(str).str.strip().map(m["products"])
            elif action == "mul_price":
                df[actual] = (pd.to_numeric(df[actual], errors="coerce") * PRICE_COEF).round(2)
            elif action == "add_qty":
                df[actual] = pd.to_numeric(df[actual], errors="coerce") + QTY_SHIFT

    # drop/rename/reorder выполняются всегда (нужны и для чистого merge без анонимизации)
    drop_cols = [
        col[k] for k in CANONICAL
        if CANONICAL[k] == "drop" and col.get(k) in df.columns
    ]
    df = df.drop(columns=drop_cols, errors="ignore")

    rename = {
        actual: cname
        for cname, actual in col.items()
        if actual in df.columns and CANONICAL.get(cname) != "drop"
    }
    df = df.rename(columns=rename)

    ordered = [c for c in DESIRED_ORDER if c in df.columns]
    extras  = [c for c in df.columns if c not in ordered]
    return df[ordered + extras]


# ── COLLECT & ARCHIVE ──────────────────────────────────────────────────────────

def collect_source_files(folder: Path) -> list[Path]:
    return sorted(
        f for f in folder.iterdir()
        if f.is_file() and f.suffix.lower() in SUPPORTED_EXTS
    )

def archive_source_files(
    files: list[Path],
    archive_folder,
    log_fn: Callable[[str], None] = print,
) -> int:
    archive_folder = Path(archive_folder)
    archive_folder.mkdir(parents=True, exist_ok=True)
    moved = 0
    for f in files:
        try:
            dest = archive_folder / f.name
            if dest.exists():
                stem, suffix, i = f.stem, f.suffix, 1
                while dest.exists():
                    dest = archive_folder / f"{stem}_{i}{suffix}"
                    i += 1
            shutil.move(str(f), dest)
            log_fn(f"  Архивирован: {f.name} → {archive_folder.name}/{dest.name}")
            moved += 1
        except Exception as e:
            log_fn(f"  ОШИБКА архивирования {f.name}: {e}")
    return moved


# ── CORE PIPELINE FUNCTIONS (importable by Dagster) ────────────────────────────

def read_and_anonymize(
    files: list[Path],
    log_fn: Callable[[str], None] = print,
    anonymize: bool = False,
    categorize: bool = False,
    name_match: bool = False,
    model: str = "qwen2.5:3b",
) -> list[tuple[str, pd.DataFrame]]:
    m = (load_mappings() if anonymize
         else {"stores": {}, "brands": {}, "companies": {}, "products": {}})
    if anonymize:
        log_fn(
            f"Маппинги загружены: магазины={len(m['stores'])}, бренды={len(m['brands'])}, "
            f"компании={len(m['companies'])}, товары={len(m['products'])}"
        )
    else:
        log_fn("ℹ️  Анонимизация выключена (anonymize=False) — данные не маскируются")

    # ── Pass 1: read all files ─────────────────────────────────────────────────
    raw_frames: list = []
    for path in files:
        try:
            log_fn(f"📖 {path.name}")
            df = read_table(path)
            df.columns = [str(c).strip() for c in df.columns]
            log_fn(f"   Строк: {len(df):,}  Столбцов: {len(df.columns)}")

            col_map = get_col_map(df)
            unmapped = [c for c in df.columns if c not in col_map.values()]
            log_fn(
                f"   Сопоставлено: {len(col_map)}  "
                f"Не определены (пройдут без изменений): {unmapped or '—'}"
            )
            raw_frames.append((path, df, col_map))

        except Exception as exc:
            log_fn(f"   ✗ ОШИБКА ({path.name}): {exc}")

    if not raw_frames:
        return []

    # ── Шаг A: категоризация (целевая система категорий) ────────────────────────
    if categorize:
        from category_matcher import apply_categories, load_target_categories
        cats = load_target_categories()
        log_fn(f"🏷️  Категоризация ({len(cats)} целевых категорий)...")
        for _, df, col in raw_frames:
            apply_categories(df, col, cats, model=model, log_fn=log_fn)

    # ── Шаг B: мэтчинг названий (авто-fuzzy, пересчитывается каждый раз) ──────────
    if name_match:
        from name_matcher import (build_name_matches, items_from_frames,
                                   save_name_matches, apply_name_matches)
        log_fn("🔗  Мэтчинг названий: анализ схожих наименований...")
        pairs   = items_from_frames(raw_frames)
        matches = build_name_matches(pairs, log_fn=log_fn)
        save_name_matches(matches)          # перезаписываем — нет накопления старых матчей
        for _, df, col in raw_frames:
            apply_name_matches(df, col, matches)

    # ── Pass 2: transform (+anon при anonymize=True) ───────────────────────────
    frames: list[tuple[str, pd.DataFrame]] = []
    for path, df, col_map in raw_frames:
        try:
            if anonymize:
                build_mappings(df, col_map, m, log_fn)
            df_out = transform(df.copy(), col_map, m, anonymize=anonymize)
            frames.append((path.name, df_out))
            log_fn(f"   ✓ Обработан: {len(df_out.columns)} столбцов")

        except Exception as exc:
            log_fn(f"   ✗ ОШИБКА ({path.name}): {exc}")

    if anonymize:
        save_mappings(m)
        log_fn(
            f"Маппинги сохранены: магазины={len(m['stores'])}, бренды={len(m['brands'])}, "
            f"компании={len(m['companies'])}, товары={len(m['products'])}"
        )
    return frames


def merge_anonymized(
    frames: list[tuple[str, pd.DataFrame]],
    log_fn: Callable[[str], None] = print,
    aggregate: bool = False,
) -> pd.DataFrame:
    file_cols = {fname: list(df.columns) for fname, df in frames}

    col_sets = [set(cols) for cols in file_cols.values()]
    common: set[str] = col_sets[0].copy()
    for s in col_sets[1:]:
        common &= s

    any_dropped = False
    for fname, cols in file_cols.items():
        dropped = [c for c in cols if c not in common]
        if dropped:
            log_fn(f"  ⚠ {fname} — удалены уникальные столбцы: {dropped}")
            any_dropped = True
    if not any_dropped:
        log_fn("  ✓ Все файлы имеют одинаковый набор столбцов")

    ordered    = [c for c in DESIRED_ORDER if c in common]
    extras     = sorted(c for c in common if c not in ordered)
    final_cols = ordered + extras
    log_fn(f"  Итоговых столбцов: {len(final_cols)} → {', '.join(final_cols)}")

    merged = pd.concat([df[final_cols] for _, df in frames], ignore_index=True)
    log_fn(f"  Строк после объединения: {len(merged):,}")

    # ── Aggregation (по умолчанию выкл — мэтчинг сохраняет все строки продаж) ───
    if aggregate:
        metric_candidates = ("qty_sold", "sales_rub", "cost_rub")
        metric_cols = [c for c in metric_candidates if c in merged.columns]
        group_cols  = [c for c in merged.columns if c not in metric_cols]

        if metric_cols and group_cols:
            for mc in metric_cols:
                numeric: pd.Series = pd.to_numeric(merged[mc], errors="coerce")
                merged[mc] = numeric.fillna(0)

            merged = (
                merged
                .groupby(group_cols, as_index=False, dropna=False)
                .agg({mc: "sum" for mc in metric_cols})
            )

            ordered2 = [c for c in DESIRED_ORDER if c in merged.columns]
            extras2  = [c for c in merged.columns if c not in ordered2]
            merged   = merged[ordered2 + extras2]

            log_fn(
                f"  Строк после агрегации: {len(merged):,}  "
                f"(схлопнуто по: {', '.join(metric_cols)})"
            )
        else:
            log_fn("  ⚠ Агрегация пропущена: нет числовых метрик или нет группирующих столбцов")

    log_fn(f"  Итого строк: {len(merged):,}")
    return merged


def save_merged(
    df: pd.DataFrame,
    output_folder,
    filename: str = "",
    log_fn: Callable[[str], None] = print,
) -> Path:
    output_folder = Path(output_folder)
    output_folder.mkdir(parents=True, exist_ok=True)
    if not filename:
        filename = f"merged_{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
    if not Path(filename).suffix:
        filename += ".xlsx"
    out = output_folder / filename
    write_xlsx(df, out)
    log_fn(f"  Сохранено → {out.resolve()}")
    return out


# ── CONFIG ─────────────────────────────────────────────────────────────────────

def load_config() -> dict:
    defaults = {
        "input_folder": "",
        "archive_folder": "",
        "output_folder": "",
        "anonymize": False,
        "categorize": False,
        "name_match": False,
        "model": "qwen2.5:3b",
    }
    if CONFIG_PATH.exists():
        try:
            existing = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            existing = {}
        return {**defaults, **existing}
    return defaults

def save_config(cfg: dict):
    existing: dict = {}
    if CONFIG_PATH.exists():
        try:
            existing = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            pass
    existing.update(cfg)
    CONFIG_PATH.write_text(
        json.dumps(existing, indent=2, ensure_ascii=False), encoding="utf-8"
    )


# ── CLI HELPERS ────────────────────────────────────────────────────────────────

def _hr(char: str = "━", n: int = 52):
    print(char * n)

def select_folder(label: str, current: str) -> str:
    cwd     = Path(".")
    folders = sorted(d for d in cwd.iterdir() if d.is_dir() and not d.name.startswith("."))
    cur_hint = f"текущая: {current}" if current else "не задана"
    print(f"\n{label}  ({cur_hint})")
    if current:
        print(f"  {'↵':>3}  оставить текущую")
    for i, d in enumerate(folders, 1):
        print(f"  {i:>3}  {d.name}/")
    if not folders and not current:
        val = input("  Нет папок. Введите путь вручную: ").strip()
        return val or "."
    while True:
        raw = input("  > ").strip()
        if not raw and current:
            return current
        if raw.isdigit() and 1 <= int(raw) <= len(folders):
            return str(folders[int(raw) - 1])
        if not raw and not current:
            print("  Текущей папки нет — выберите из списка.")
        else:
            print(f"  Введите номер 1–{len(folders)} или Enter.")


# ── MAIN ───────────────────────────────────────────────────────────────────────

def main():
    print()
    _hr()
    print("  🔀  Retail Pipeline — объединение таблиц продаж")
    _hr()

    cfg = load_config()

    # 1. Input folder
    cfg["input_folder"] = select_folder(
        "Папка с исходными файлами", cfg["input_folder"]
    )
    save_config(cfg)

    input_dir = Path(cfg["input_folder"])
    files = collect_source_files(input_dir)
    if not files:
        print(f"\n  ⚠  Нет поддерживаемых файлов в {input_dir}/")
        print(f"     Форматы: {', '.join(sorted(SUPPORTED_EXTS))}")
        sys.exit(1)

    print(f"\n  Найдено файлов: {len(files)}")
    for i, f in enumerate(files, 1):
        kb = max(f.stat().st_size // 1024, 1)
        print(f"  {i:>3}  {f.name}  ({kb} КБ)")

    # 2. Archive folder (for processed source files)
    cfg["archive_folder"] = select_folder(
        "Папка для архива обработанных исходников", cfg["archive_folder"]
    )
    save_config(cfg)

    # 3. Output folder
    cfg["output_folder"] = select_folder(
        "Папка для сохранения результата", cfg["output_folder"]
    )
    save_config(cfg)

    # 4. Output filename
    default_name = f"merged_{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
    print(f"\nИмя файла результата  [Enter → {default_name}]:")
    raw_name = input("  > ").strip() or default_name
    if not Path(raw_name).suffix:
        raw_name += ".xlsx"

    # ── Pipeline ───────────────────────────────────────────────────────────────
    print()
    _hr("─")
    t0 = time.time()

    def log(msg: str):
        print(f"  [{time.time() - t0:5.1f}s] {msg}")

    anon = cfg.get("anonymize", False)
    log(f"📖  Чтение и обработка {len(files)} файлов (анонимизация: {'вкл' if anon else 'выкл'})...")
    frames = read_and_anonymize(
        files,
        log_fn=log,
        anonymize=anon,
        categorize=cfg.get("categorize", False),
        name_match=cfg.get("name_match", False),
        model=cfg.get("model", "qwen2.5:3b"),
    )

    if not frames:
        log("❌  Ни один файл не прочитан. Выход.")
        sys.exit(1)

    log("🔀  Объединение таблиц...")
    merged = merge_anonymized(frames, log_fn=log)

    log("💾  Сохранение...")
    out_path = save_merged(merged, cfg["output_folder"], raw_name, log_fn=log)

    log(f"📦  Архивирование исходников → {cfg['archive_folder']}/")
    n_archived = archive_source_files(files, cfg["archive_folder"], log_fn=log)

    elapsed = time.time() - t0
    _hr("─")
    print()
    print("  📊  Итог по файлам:")
    for fname, df in frames:
        print(f"       {fname:<46} {len(df):>8,} строк")
    errors = [f.name for f in files if f.name not in {n for n, _ in frames}]
    for fname in errors:
        print(f"       {fname:<46}  — ошибка чтения")
    print()
    print(f"  Обработано файлов:   {len(frames)}")
    print(f"  Итого строк:         {len(merged):,}")
    print(f"  Архивировано:        {n_archived} файлов → {cfg['archive_folder']}/")
    print()
    _hr()
    print(f"\n  ✅  Готово за {elapsed:.2f} сек")
    print(f"  →  {out_path.resolve()}\n")


if __name__ == "__main__":
    main()